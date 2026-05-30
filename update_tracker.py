import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\raghu\VV\VardhiniVastu_SEO_Tracker_v6.xlsx')
ws = wb['Pages Upload Pipeline']

updates = {
    3: {
        'status': 'Done',
        'wp_ids': '360,351,352,353,356,357,358,342,343,344,359,348,350,349,354,355,361,338,341,340,339,345,346,347'
    },
    6: {
        'status': 'Done',
        'wp_ids': '118,98,931,932,108,110,94,120,937,100,116,112,941,942,96'
    },
    7: {
        'status': 'Done',
        'wp_ids': '262,257,259,264,255,258,265,261,263,256,260,254,280,281,282'
    },
    10: {
        'status': 'Done',
        'wp_ids': '920,921,157,148,150,142,146,927,928'
    },
    11: {
        'status': 'Done',
        'wp_ids': '183,187,179,185,181,177,189,191'
    },
}

# Print current headers to confirm column positions
headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
print('Headers:', headers)

for row_num, data in updates.items():
    folder = ws.cell(row=row_num, column=2).value
    ws.cell(row=row_num, column=5).value = data['status']  # Status col E
    ws.cell(row=row_num, column=6).value = data['wp_ids']  # WP IDs col F
    print(f'Row {row_num} ({folder}): Status=Done, WP IDs set')

wb.save(r'C:\Users\raghu\VV\VardhiniVastu_SEO_Tracker_v6.xlsx')
print('Saved.')
