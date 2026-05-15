# VardhiniVastu.in - SEO Master Tracker Builder
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ── Colour palette ──────────────────────────────────────────────────────────
NAVY        = "1E3A5F"
GOLD        = "B8860B"
GREEN       = "2D6A4F"
AMBER       = "D4740E"
RED         = "C53030"
CREAM       = "FAF9F7"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GREEN = "D5ECD8"
LIGHT_RED   = "F9DEDE"
LIGHT_AMBER = "FAF0DC"
WHITE       = "FFFFFF"
GRAY        = "E8E8E8"
MID_GRAY    = "AAAAAA"

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Task data ────────────────────────────────────────────────────────────────
# Columns: Category | Task | Priority | Status | Notes/URL | Target | Done Date
DONE     = "✅ Done"
PENDING  = "⏳ Pending"
PROGRESS = "🔄 In Progress"

tasks = [

  # ── 1. SEO Strategy & Competitor Analysis ──────────────────────────────
  ("🎯 Strategy & Competitors", "Identify top 5 Vastu consultant competitors in Bangalore",
   "High", DONE, "Competitors: Mystic Space Vastu, VastuPlus, Vastu Devayah Namah, Astro Vastu, local consultants", "", ""),

  ("🎯 Strategy & Competitors", "Map competitor keyword universe (city pages, locality pages, room pages)",
   "High", DONE, "Locality + city + room page structure benchmarked against competitors", "", ""),

  ("🎯 Strategy & Competitors", "Define SEO content roadmap: Rooms, Localities, Cities, International, Educational, Directions",
   "High", DONE, "Full roadmap built — 68+ pages planned and executed", "", ""),

  ("🎯 Strategy & Competitors", "Rank Math 90+ blueprint: title format, keyword density, H2 rule, image alt, external link",
   "High", DONE, "Blueprint documented in memory/feedback_rankmath_blueprint.md", "", ""),

  ("🎯 Strategy & Competitors", "Set focus keywords per page (Rank Math / meta tags)",
   "High", PENDING, "Each page has SEO title + meta desc in HTML comment — needs to be set in Rank Math for all 68 pages", "", ""),

  ("🎯 Strategy & Competitors", "Track rankings for top 20 target keywords monthly",
   "Medium", PENDING, "Use Google Search Console + Rank Math rank tracker", "", ""),

  ("🎯 Strategy & Competitors", "Competitor backlink gap analysis",
   "Medium", PENDING, "Run /seo backlinks on top 3 competitors", "", ""),

  # ── 2. Technical SEO ────────────────────────────────────────────────────
  ("🔧 Technical SEO", "vv-blank.php custom template (serves standalone HTML without theme wrapper)",
   "High", DONE, "/wp-content/themes/astra/vv-blank.php created and verified", "", ""),

  ("🔧 Technical SEO", "WP permalink structure set to /%postname%/",
   "High", DONE, "Clean URL slugs confirmed on all pages", "", ""),

  ("🔧 Technical SEO", "XML sitemap generation & submission to Google Search Console",
   "High", PENDING, "Rank Math auto-generates sitemap — verify it includes all 68+ new pages, then submit in GSC", "", ""),

  ("🔧 Technical SEO", "robots.txt — verify crawl rules, no important pages blocked",
   "High", PENDING, "Check vardhinivastu.in/robots.txt — ensure locality/city/room pages are crawlable", "", ""),

  ("🔧 Technical SEO", "SSL / HTTPS — verify all pages load over HTTPS",
   "High", PENDING, "Check for mixed content warnings", "", ""),

  ("🔧 Technical SEO", "Core Web Vitals (LCP, CLS, INP) — baseline measurement",
   "High", PENDING, "Run PageSpeed Insights on homepage + 3 locality pages", "", ""),

  ("🔧 Technical SEO", "Mobile rendering — all 68 pages render correctly on mobile",
   "High", PENDING, "Test Astra theme wrapper on fragment pages vs standalone pages", "", ""),

  ("🔧 Technical SEO", "Canonical tags — no duplicate content across city/locality variations",
   "Medium", PENDING, "Check Rank Math canonical settings on all pages", "", ""),

  ("🔧 Technical SEO", "Internal linking audit — link locality pages to city pages, room pages to consultation CTA",
   "Medium", PENDING, "Build cross-link structure: rooms → localities → cities → homepage", "", ""),

  ("🔧 Technical SEO", "Image optimisation — WebP conversion + alt text on all pages",
   "Medium", PENDING, "Upload actual images to /wp-content/uploads/ for all page placeholder img tags", "", ""),

  ("🔧 Technical SEO", "404 error audit — fix any broken internal links",
   "Low", PENDING, "Scan with Screaming Frog or GSC Coverage report after indexing", "", ""),

  # ── 3. Content: Residential Rooms (25 pages) ────────────────────────────
  ("📄 Residential Rooms", "vastu-for-kitchen",          "High", DONE, "vardhinivastu.in/vastu-for-kitchen · WP ID 283",         "", ""),
  ("📄 Residential Rooms", "vastu-for-bedroom",          "High", DONE, "vardhinivastu.in/vastu-for-bedroom · WP ID 284",         "", ""),
  ("📄 Residential Rooms", "vastu-for-living-room",      "High", DONE, "vardhinivastu.in/vastu-for-living-room · WP ID 285",     "", ""),
  ("📄 Residential Rooms", "vastu-for-dining-room",      "High", DONE, "vardhinivastu.in/vastu-for-dining-room · WP ID 286",     "", ""),
  ("📄 Residential Rooms", "vastu-for-bathroom",         "High", DONE, "vardhinivastu.in/vastu-for-bathroom · WP ID 287",        "", ""),
  ("📄 Residential Rooms", "vastu-for-pooja-room",       "High", DONE, "vardhinivastu.in/vastu-for-pooja-room · WP ID 288",      "", ""),
  ("📄 Residential Rooms", "vastu-for-main-entrance",    "High", DONE, "vardhinivastu.in/vastu-for-main-entrance · WP ID 289",   "", ""),
  ("📄 Residential Rooms", "vastu-for-study-room",       "High", DONE, "vardhinivastu.in/vastu-for-study-room · WP ID 290",      "", ""),
  ("📄 Residential Rooms", "vastu-for-guest-room",       "Medium", DONE, "vardhinivastu.in/vastu-for-guest-room · WP ID 291",    "", ""),
  ("📄 Residential Rooms", "vastu-for-childrens-room",   "High", DONE, "vardhinivastu.in/vastu-for-childrens-room · WP ID 292", "", ""),
  ("📄 Residential Rooms", "vastu-for-master-bedroom",   "High", DONE, "vardhinivastu.in/vastu-for-master-bedroom · WP ID 293", "", ""),
  ("📄 Residential Rooms", "vastu-for-mandir-at-home",   "High", DONE, "vardhinivastu.in/vastu-for-mandir-at-home · WP ID 294", "", ""),
  ("📄 Residential Rooms", "vastu-for-balcony",          "Medium", DONE, "vardhinivastu.in/vastu-for-balcony · WP ID 295",      "", ""),
  ("📄 Residential Rooms", "vastu-for-mirrors",          "Medium", DONE, "vardhinivastu.in/vastu-for-mirrors · WP ID 296",       "", ""),
  ("📄 Residential Rooms", "vastu-for-shoe-rack",        "Low", DONE, "vardhinivastu.in/vastu-for-shoe-rack · WP ID 297",        "", ""),
  ("📄 Residential Rooms", "vastu-for-staircase",        "Medium", DONE, "vardhinivastu.in/vastu-for-staircase · WP ID 298",     "", ""),
  ("📄 Residential Rooms", "vastu-for-basement",         "Medium", DONE, "vardhinivastu.in/vastu-for-basement · WP ID 299",      "", ""),
  ("📄 Residential Rooms", "vastu-for-car-parking",      "Medium", DONE, "vardhinivastu.in/vastu-for-car-parking · WP ID 300",   "", ""),
  ("📄 Residential Rooms", "vastu-for-garden",           "Medium", DONE, "vardhinivastu.in/vastu-for-garden · WP ID 301",        "", ""),
  ("📄 Residential Rooms", "vastu-for-water-tank",       "Medium", DONE, "vardhinivastu.in/vastu-for-water-tank · WP ID 302",    "", ""),
  ("📄 Residential Rooms", "vastu-for-swimming-pool",    "Low", DONE, "vardhinivastu.in/vastu-for-swimming-pool · WP ID 303",    "", ""),
  ("📄 Residential Rooms", "vastu-for-septic-tank",      "Low", DONE, "vardhinivastu.in/vastu-for-septic-tank · WP ID 304",      "", ""),
  ("📄 Residential Rooms", "vastu-for-wall-clock",       "Low", DONE, "vardhinivastu.in/vastu-for-wall-clock · WP ID 305",       "", ""),
  ("📄 Residential Rooms", "vastu-for-colors",           "High", DONE, "vardhinivastu.in/vastu-for-colors · WP ID 306",          "", ""),
  ("📄 Residential Rooms", "vastu-for-toilet",           "Medium", DONE, "vardhinivastu.in/vastu-for-toilet · WP ID 307",        "", ""),

  # ── 4. Content: Bangalore Localities (20 pages) ─────────────────────────
  ("📍 Bangalore Localities", "vastu-consultant-whitefield",      "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-indiranagar",     "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-koramangala",     "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-hsr-layout",      "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-electronic-city", "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-jp-nagar",        "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-jayanagar",       "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-bannerghatta-road","Medium", DONE, "WP ID 267", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-sarjapur-road",   "Medium", DONE, "WP ID 268", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-marathahalli",    "Medium", DONE, "WP ID 269", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-yelahanka",       "Medium", DONE, "WP ID 270", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-hebbal",          "Medium", DONE, "WP ID 271", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-rajajinagar",     "Medium", DONE, "WP ID 272", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-malleswaram",     "Medium", DONE, "WP ID 273", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-basavanagudi",    "Medium", DONE, "WP ID 274", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-btm-layout",      "Medium", DONE, "WP ID 275", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-banaswadi",       "Low", DONE, "WP ID 276", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-bellandur",       "Medium", DONE, "WP ID 277", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-kr-puram",        "Low", DONE, "WP ID 278", "", ""),
  ("📍 Bangalore Localities", "vastu-consultant-nagarbhavi",      "Low", DONE, "WP ID 279", "", ""),

  # ── 5. Content: Indian Cities (15 pages) ────────────────────────────────
  ("🏙️ Indian Cities", "vastu-consultant-chennai",     "High", DONE, "WP ID 255", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-mumbai",      "High", DONE, "WP ID 263", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-delhi",       "High", DONE, "WP ID 258", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-hyderabad",   "High", DONE, "WP ID 260", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-pune",        "High", DONE, "WP ID 265", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-ahmedabad",   "Medium", DONE, "WP ID 254", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-kolkata",     "Medium", DONE, "WP ID 261", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-gurugram",    "Medium", DONE, "WP ID 259", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-noida",       "Medium", DONE, "WP ID 264", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-coimbatore",  "Medium", DONE, "WP ID 256", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-mysuru",      "Medium", DONE, "WP ID 262", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-mangaluru",   "Medium", DONE, "WP ID 261 area", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-kochi",       "Medium", DONE, "WP ID 280", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-jaipur",      "Medium", DONE, "WP ID 281", "", ""),
  ("🏙️ Indian Cities", "vastu-consultant-chandigarh",  "Low", DONE, "WP ID 282", "", ""),

  # ── 6. Content: International Locations (8 pages) ───────────────────────
  ("🌍 International", "vastu-consultant-usa",         "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("🌍 International", "vastu-consultant-uk",          "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("🌍 International", "vastu-consultant-canada",      "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("🌍 International", "vastu-consultant-australia",   "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("🌍 International", "vastu-consultant-uae",         "High", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("🌍 International", "vastu-consultant-singapore",   "Medium", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("🌍 International", "vastu-consultant-new-zealand", "Low", DONE, "WP live — standalone HTML via vv-blank", "", ""),
  ("🌍 International", "vastu-consultant-germany",     "Low", DONE, "WP live — standalone HTML via vv-blank", "", ""),

  # ── 7. Content: Educational Hub (15 pages — old format, needs rewrite) ──
  ("📚 Educational Hub", "Rewrite 15 Educational Hub pages from old <article class='vastu-page'> format to new vv-page-hero format",
   "Medium", PENDING, "Files exist locally in pages/educational-hub/ — content good, template needs upgrading", "", ""),
  ("📚 Educational Hub", "vastu-shastra-history",  "Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-shastra-principles","Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-five-elements",     "Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-16-zones",          "Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-wealth",        "High", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-health",        "High", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-marriage",      "High", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-career",        "High", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-remedies",          "High", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-dosha",             "Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-flat",          "Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-office",        "Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-plot",          "Medium", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vastu-for-new-home",      "High", PENDING, "Needs rewrite to vv-page-hero format then push to WP", "", ""),
  ("📚 Educational Hub", "vids-system-explainer",   "High", PENDING, "VIDS™ unique selling page — needs to be built fresh", "", ""),

  # ── 8. Content: House Directions (10 pages — old format) ────────────────
  ("🧭 House Directions", "Rewrite 10 House Direction pages from old format to vv-page-hero then push to WP",
   "Medium", PENDING, "Covers: North-facing, South-facing, East-facing, West-facing, NE, NW, SE, SW + corner plots", "", ""),
  ("🧭 House Directions", "vastu-for-north-facing-house", "High", PENDING, "High search volume — priority rewrite", "", ""),
  ("🧭 House Directions", "vastu-for-south-facing-house", "High", PENDING, "High search volume (controversial topic — good for SEO)", "", ""),
  ("🧭 House Directions", "vastu-for-east-facing-house",  "High", PENDING, "High search volume", "", ""),
  ("🧭 House Directions", "vastu-for-west-facing-house",  "Medium", PENDING, "", "", ""),
  ("🧭 House Directions", "vastu-for-northeast-facing",   "Medium", PENDING, "", "", ""),
  ("🧭 House Directions", "vastu-for-northwest-facing",   "Medium", PENDING, "", "", ""),
  ("🧭 House Directions", "vastu-for-southeast-facing",   "Medium", PENDING, "", "", ""),
  ("🧭 House Directions", "vastu-for-southwest-facing",   "Medium", PENDING, "SW is the most feared — high search intent", "", ""),
  ("🧭 House Directions", "vastu-for-corner-plot",        "Medium", PENDING, "Corner plot vastu — niche but high intent", "", ""),

  # ── 9. Other Core Pages ─────────────────────────────────────────────────
  ("📝 Core Pages", "Homepage — SEO-optimise title, meta, H1, above-fold content, schema",
   "High", PENDING, "Homepage is the primary landing page — needs full Rank Math 90+ treatment", "", ""),
  ("📝 Core Pages", "About / Raghavendra Hebbur — E-E-A-T author page with credentials, photo, LinkedIn",
   "High", PENDING, "Crucial for E-E-A-T — named expert signals boost all page rankings", "", ""),
  ("📝 Core Pages", "Contact / Consultation booking page",
   "High", PENDING, "Needs optimised CTA, schema, Google Maps embed, NAP", "", ""),
  ("📝 Core Pages", "Testimonials / Reviews page",
   "High", PENDING, "Aggregate reviews with Review schema markup", "", ""),
  ("📝 Core Pages", "Services overview page (what is included in a VIDS™ consultation)",
   "High", PENDING, "Service schema + pricing info page", "", ""),
  ("📝 Core Pages", "Blog — launch with 4 cornerstone posts targeting informational queries",
   "Medium", PENDING, "Topic ideas: vastu mistakes, vastu vs feng shui, vastu for 2024, apartment vastu", "", ""),
  ("📝 Core Pages", "FAQ hub page — aggregate all FAQ schema in one place",
   "Low", PENDING, "Useful for AI Overviews / featured snippets", "", ""),

  # ── 10. Local SEO & GMB ──────────────────────────────────────────────────
  ("📌 GMB & Local SEO", "Claim / verify Google Business Profile (GMB)",
   "High", PENDING, "Business name: Vardhini Vastu — verify at business.google.com", "", ""),
  ("📌 GMB & Local SEO", "GMB: Set primary category = 'Vastu Consultant'",
   "High", PENDING, "Primary: Vastu Consultant; Secondary: Feng Shui Consultant, Spiritual Counselor", "", ""),
  ("📌 GMB & Local SEO", "GMB: Complete business description (750 chars max) with focus keywords",
   "High", PENDING, "Include: VIDS™, Raghavendra Hebbur, Bangalore, zero-demolition, 16-zone analysis", "", ""),
  ("📌 GMB & Local SEO", "GMB: Add service area — cover all 20 Bangalore localities + Tier-1 cities",
   "High", PENDING, "Add each locality as a service area in GMB settings", "", ""),
  ("📌 GMB & Local SEO", "GMB: Upload minimum 10 high-quality photos (exterior, consultation, before/after)",
   "High", PENDING, "Cover photo, logo, team photo, work samples", "", ""),
  ("📌 GMB & Local SEO", "GMB: Add all services with descriptions and prices (consultation types)",
   "High", PENDING, "Residential, Commercial, Plot analysis, Online consultation", "", ""),
  ("📌 GMB & Local SEO", "GMB: Add Q&A — pre-populate with 10 common vastu questions",
   "Medium", PENDING, "Questions should match FAQ content on pages", "", ""),
  ("📌 GMB & Local SEO", "GMB: Enable messaging + set auto-reply",
   "Medium", PENDING, "WhatsApp is primary — also enable GMB messaging", "", ""),
  ("📌 GMB & Local SEO", "GMB: Set up weekly GMB posts (tip, offer, or event)",
   "Medium", PENDING, "1 post/week minimum — repurpose room/direction page content", "", ""),
  ("📌 GMB & Local SEO", "GMB: Google Reviews — get first 10 reviews from past clients",
   "High", PENDING, "Target: 10+ reviews, 4.8+ rating in first 90 days", "", ""),
  ("📌 GMB & Local SEO", "GMB: Respond to ALL reviews within 24 hours",
   "High", PENDING, "Ongoing — positive and negative both need responses", "", ""),
  ("📌 GMB & Local SEO", "GMB: Add booking link / appointment URL",
   "High", PENDING, "Link to /vastu-consultant-bangalore-contact/", "", ""),
  ("📌 GMB & Local SEO", "NAP consistency — Name, Address, Phone identical across website, GMB, citations",
   "High", PENDING, "NAP: Vardhini Vastu | Bangalore, Karnataka | +91 97391 05574", "", ""),
  ("📌 GMB & Local SEO", "Citation building — Tier 1: Justdial, Sulekha, IndiaMart, UrbanClap",
   "High", PENDING, "Submit to Justdial, Sulekha, UrbanClap/Urban Company, NoBroker, 99acres", "", ""),
  ("📌 GMB & Local SEO", "Citation building — Tier 2: local Bangalore directories, housing.com, magicbricks",
   "Medium", PENDING, "Housing.com consultant directory, Magicbricks vastu section", "", ""),
  ("📌 GMB & Local SEO", "LocalBusiness schema on homepage and contact page",
   "High", PENDING, "Already in room/locality pages — add to homepage and contact page too", "", ""),
  ("📌 GMB & Local SEO", "Bing Places listing (mirrors GMB — quick win)",
   "Low", PENDING, "Import from GMB at bingplaces.com", "", ""),
  ("📌 GMB & Local SEO", "Apple Maps listing (growing in India with iPhone adoption)",
   "Low", PENDING, "Register at mapsconnect.apple.com", "", ""),

  # ── 11. Off-Page & Authority ─────────────────────────────────────────────
  ("🔗 Off-Page & Authority", "Get listed in Vastu/Astrology expert directories (Astroyogi, Ganeshaspeaks)",
   "Medium", PENDING, "", "", ""),
  ("🔗 Off-Page & Authority", "Guest article on housing/real estate blog (housing.com, 99acres, magicbricks blog)",
   "High", PENDING, "Target: 1 guest post/month with link back to vardhinivastu.in", "", ""),
  ("🔗 Off-Page & Authority", "YouTube channel — 2 short-form vastu tip videos/month",
   "Medium", PENDING, "Short explainer videos rank well for vastu queries — embed on pages", "", ""),
  ("🔗 Off-Page & Authority", "LinkedIn presence — Raghavendra Hebbur profile optimised",
   "High", PENDING, "Author credibility signal for E-E-A-T", "", ""),
  ("🔗 Off-Page & Authority", "PR / media coverage — approach home décor / lifestyle journalists",
   "Low", PENDING, "Target: 1 press mention/quarter in a home/lifestyle publication", "", ""),

  # ── 12. Analytics & Tracking ────────────────────────────────────────────
  ("📊 Analytics", "Google Search Console — verify site, submit sitemap, monitor Coverage",
   "High", PENDING, "Submit sitemap.xml — monitor for crawl errors on all 68+ new pages", "", ""),
  ("📊 Analytics", "Google Analytics 4 — verify tracking on all pages",
   "High", PENDING, "Check GA4 realtime when a page is visited — confirm event firing", "", ""),
  ("📊 Analytics", "Rank Math Analytics — connect GSC to Rank Math dashboard",
   "High", PENDING, "Enables keyword rank tracking inside WP admin", "", ""),
  ("📊 Analytics", "Set up conversion tracking — WhatsApp click, contact form submission, call click",
   "High", PENDING, "GA4 events for all 3 CTA types on all pages", "", ""),
  ("📊 Analytics", "Monthly SEO report — rankings, traffic, GMB views, calls",
   "Medium", PENDING, "Baseline report at 30 days, then monthly", "", ""),
]

# ── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Dashboard ───────────────────────────────────────────────────────
ws_dash = wb.active
ws_dash.title = "📊 Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.column_dimensions["A"].width = 35
ws_dash.column_dimensions["B"].width = 18
ws_dash.column_dimensions["C"].width = 18
ws_dash.column_dimensions["D"].width = 35
ws_dash.row_dimensions[1].height = 60

# Title row
ws_dash.merge_cells("A1:D1")
c = ws_dash["A1"]
c.value = "Vardhini Vastu — SEO Master Tracker"
c.fill = fill(NAVY)
c.font = Font(bold=True, color=WHITE, size=20, name="Calibri")
c.alignment = centre()

ws_dash.merge_cells("A2:D2")
c = ws_dash["A2"]
c.value = f"vardhinivastu.in  ·  Raghavendra Hebbur  ·  Last updated: {date.today().strftime('%d %b %Y')}"
c.fill = fill(GOLD)
c.font = Font(bold=False, color=WHITE, size=12, name="Calibri")
c.alignment = centre()

# Stats
total   = len(tasks)
done    = sum(1 for t in tasks if t[3] == DONE)
pending = sum(1 for t in tasks if t[3] == PENDING)
inprog  = sum(1 for t in tasks if t[3] == PROGRESS)
pct     = int(done / total * 100)

ws_dash.row_dimensions[4].height = 14
ws_dash["A4"].value = ""; ws_dash["A4"].fill = fill(CREAM)

headers = ["Metric", "Count", "% of Total", "Description"]
for col, h in enumerate(headers, 1):
    c = ws_dash.cell(row=5, column=col, value=h)
    c.fill = fill(NAVY); c.font = font(bold=True); c.alignment = centre()
    c.border = border()

stats = [
    ("Total Tasks",        total,   "100%",          "All tasks across all categories"),
    ("✅ Completed",       done,    f"{pct}%",        "Done — live on site or process complete"),
    ("⏳ Pending",         pending, f"{100-pct}%",    "Not yet started"),
    ("🔄 In Progress",     inprog,  "—",              "Currently being worked on"),
    ("Pages Live on WP",   68,      "—",              "Rooms 25 + Localities 20 + Cities 15 + International 8"),
    ("GMB Tasks",          18,      "—",              "All pending — highest priority local SEO lever"),
]
for r, (label, count, pct_s, desc) in enumerate(stats, 6):
    ws_dash.row_dimensions[r].height = 22
    row_fill = fill(LIGHT_GREEN) if "Completed" in label else fill(LIGHT_RED) if "Pending" in label else fill(LIGHT_BLUE)
    for col, val in enumerate([label, count, pct_s, desc], 1):
        c = ws_dash.cell(row=r, column=col, value=val)
        c.fill = row_fill
        c.font = Font(bold=(col == 1), color="111111", size=11, name="Calibri")
        c.alignment = centre() if col in (2, 3) else left()
        c.border = border()

# Legend
ws_dash.row_dimensions[13].height = 14
ws_dash["A13"].value = ""; ws_dash["A13"].fill = fill(CREAM)

ws_dash.merge_cells("A14:D14")
c = ws_dash["A14"]
c.value = "STATUS LEGEND"
c.fill = fill(NAVY); c.font = font(bold=True, size=12); c.alignment = centre()

legend = [
    ("✅ Done",         "Task complete, page live on site",                  LIGHT_GREEN),
    ("⏳ Pending",      "Not started — needs to be picked up",               LIGHT_RED),
    ("🔄 In Progress",  "Currently being worked on",                         LIGHT_AMBER),
]
for r, (status, desc, bg) in enumerate(legend, 15):
    ws_dash.row_dimensions[r].height = 22
    for col, val in enumerate([status, "", desc, ""], 1):
        c = ws_dash.cell(row=r, column=col, value=val)
        c.fill = fill(bg)
        c.font = Font(color="111111", size=11, name="Calibri")
        c.alignment = left(); c.border = border()
    ws_dash.merge_cells(f"C{r}:D{r}")

# Priority legend
ws_dash.row_dimensions[19].height = 14
ws_dash["A19"].value = ""; ws_dash["A19"].fill = fill(CREAM)

ws_dash.merge_cells("A20:D20")
c = ws_dash["A20"]
c.value = "PRIORITY LEGEND"
c.fill = fill(NAVY); c.font = font(bold=True, size=12); c.alignment = centre()

plegend = [
    ("🔴 High",   "Critical for rankings or conversions — do first",   "FFDEDE"),
    ("🟡 Medium", "Important optimisation — do within 30 days",        "FFF9DC"),
    ("🟢 Low",    "Nice to have — backlog item",                       "E2F5E2"),
]
for r, (p, desc, bg) in enumerate(plegend, 21):
    ws_dash.row_dimensions[r].height = 22
    for col, val in enumerate([p, "", desc, ""], 1):
        c = ws_dash.cell(row=r, column=col, value=val)
        c.fill = fill(bg)
        c.font = Font(color="111111", size=11, name="Calibri")
        c.alignment = left(); c.border = border()
    ws_dash.merge_cells(f"C{r}:D{r}")

# ── Sheet 2: Master Task List ────────────────────────────────────────────────
ws = wb.create_sheet("✅ Master Task List")
ws.sheet_view.showGridLines = False

# Column widths
col_widths = [5, 28, 48, 10, 18, 48, 15, 15]
col_letters = [get_column_letter(i+1) for i in range(len(col_widths))]
for i, w in enumerate(col_widths):
    ws.column_dimensions[col_letters[i]].width = w

# Header row
ws.row_dimensions[1].height = 50
headers = ["#", "Category", "Task / Page", "Priority", "Status", "Notes / URL", "Target Date", "Done Date"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = font(bold=True, size=11)
    c.alignment = centre()
    c.border = border()

# Data validation for Status column
dv_status = DataValidation(
    type="list",
    formula1='"✅ Done,⏳ Pending,🔄 In Progress"',
    allow_blank=False
)
ws.add_data_validation(dv_status)

dv_priority = DataValidation(
    type="list",
    formula1='"High,Medium,Low"',
    allow_blank=False
)
ws.add_data_validation(dv_priority)

# Category colour map
cat_colours = {
    "🎯 Strategy & Competitors": ("1A3A6B", "D6E4F7"),
    "🔧 Technical SEO":          ("1B4332", "D5ECD8"),
    "📄 Residential Rooms":      ("5C3317", "FDEBD0"),
    "📍 Bangalore Localities":   ("4A235A", "EAD5F5"),
    "🏙️ Indian Cities":          ("145A32", "D5F5E3"),
    "🌍 International":          ("154360", "D6EAF8"),
    "📚 Educational Hub":        ("784212", "FAE5D3"),
    "🧭 House Directions":       ("1B2631", "D5D8DC"),
    "📝 Core Pages":             ("212F3D", "D6DBDF"),
    "📌 GMB & Local SEO":        ("C0392B", "FADBD8"),
    "🔗 Off-Page & Authority":   ("1A5276", "D6EAF8"),
    "📊 Analytics":              ("117A65", "D1F2EB"),
}

status_bg = {
    DONE:     LIGHT_GREEN,
    PENDING:  LIGHT_RED,
    PROGRESS: LIGHT_AMBER,
}
priority_bg = {
    "High":   "FFDEDE",
    "Medium": "FFF9DC",
    "Low":    "E2F5E2",
}

row_num = 2
prev_cat = None
for idx, (cat, task, priority, status, notes, target, done_dt) in enumerate(tasks, 1):
    # Category separator row
    if cat != prev_cat:
        if prev_cat is not None:
            ws.row_dimensions[row_num].height = 8
            for col in range(1, 9):
                c = ws.cell(row=row_num, column=col, value="")
                c.fill = fill(CREAM)
            row_num += 1
        ws.row_dimensions[row_num].height = 28
        ws.merge_cells(f"A{row_num}:H{row_num}")
        hdr_col, _ = cat_colours.get(cat, (NAVY, WHITE))
        c = ws.cell(row=row_num, column=1, value=f"  {cat}")
        c.fill = fill(hdr_col)
        c.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        row_num += 1
        prev_cat = cat

    ws.row_dimensions[row_num].height = 38
    row_data = [idx, cat, task, priority, status, notes, target, done_dt]
    _, bg = cat_colours.get(cat, (NAVY, "F5F5F5"))

    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        if col == 4:   # Priority
            c.fill = fill(priority_bg.get(priority, WHITE))
        elif col == 5:  # Status
            c.fill = fill(status_bg.get(status, WHITE))
        else:
            c.fill = fill(bg)
        c.font = Font(
            color="111111", size=10, name="Calibri",
            bold=(col in (2, 3))
        )
        c.alignment = centre() if col in (1, 4, 5) else left()
        c.border = border()

    # Register validation
    dv_status.add(ws.cell(row=row_num, column=5))
    dv_priority.add(ws.cell(row=row_num, column=4))

    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# ── Sheet 3: GMB Quick-Start Checklist ───────────────────────────────────────
ws_gmb = wb.create_sheet("📌 GMB Quick-Start")
ws_gmb.sheet_view.showGridLines = False
ws_gmb.column_dimensions["A"].width = 5
ws_gmb.column_dimensions["B"].width = 40
ws_gmb.column_dimensions["C"].width = 45
ws_gmb.column_dimensions["D"].width = 18
ws_gmb.column_dimensions["E"].width = 18

ws_gmb.merge_cells("A1:E1")
c = ws_gmb["A1"]
c.value = "Google Business Profile (GMB) — Quick-Start Checklist"
c.fill = fill("C53030"); c.font = font(bold=True, size=16); c.alignment = centre()
ws_gmb.row_dimensions[1].height = 45

ws_gmb.merge_cells("A2:E2")
c = ws_gmb["A2"]
c.value = "Complete in this order — GMB is the #1 local ranking lever for Vastu consultants in Bangalore"
c.fill = fill(AMBER); c.font = font(size=11); c.alignment = centre()
ws_gmb.row_dimensions[2].height = 25

gmb_headers = ["#", "Task", "What to Write / Do", "Status", "Done Date"]
for col, h in enumerate(gmb_headers, 1):
    c = ws_gmb.cell(row=3, column=col, value=h)
    c.fill = fill(NAVY); c.font = font(bold=True); c.alignment = centre(); c.border = border()
ws_gmb.row_dimensions[3].height = 28

gmb_tasks = [
    (1,  "Claim & verify GMB listing",
     "Go to business.google.com → Add your business → Verify by postcard or phone",
     PENDING),
    (2,  "Business Name",
     "Vardhini Vastu  (exactly — no extra keywords)",
     PENDING),
    (3,  "Primary Category",
     "Vastu Consultant",
     PENDING),
    (4,  "Additional Categories",
     "Feng Shui Consultant | Spiritual Counselor | Home Consultant",
     PENDING),
    (5,  "Business Description (750 chars)",
     "Vardhini Vastu offers professional Vastu Shastra consultancy by Raghavendra Hebbur in Bangalore. Specialising in the VIDS™ (Vardhini Integrated Direction System) — a 16-zone, degree-accurate Vastu analysis with zero-demolition corrections. Services: residential vastu, commercial vastu, plot analysis, and online consultations across India and worldwide.",
     PENDING),
    (6,  "Phone Number",
     "+91 97391 05574  (must match website and citations exactly)",
     PENDING),
    (7,  "Website URL",
     "https://vardhinivastu.in",
     PENDING),
    (8,  "Address",
     "Add full Bangalore address (or hide address if home-based + set service area)",
     PENDING),
    (9,  "Service Area",
     "Add all 20 Bangalore localities + Hyderabad, Chennai, Mumbai, Delhi, Pune (and other cities you serve)",
     PENDING),
    (10, "Opening Hours",
     "Mon–Sat 10am–7pm / Sun By Appointment  (set accurate hours — affects call rate)",
     PENDING),
    (11, "Services — add each with description",
     "1) Residential Vastu Consultation 2) Commercial Vastu 3) Online Vastu Consultation 4) Plot/Land Vastu 5) VIDS™ Analysis",
     PENDING),
    (12, "Appointment / Booking Link",
     "https://vardhinivastu.in/vastu-consultant-bangalore-contact/",
     PENDING),
    (13, "Upload Cover Photo",
     "Professional photo of Raghavendra Hebbur OR clean vastu compass image (1332×750px)",
     PENDING),
    (14, "Upload Logo",
     "Vardhini Vastu logo (250×250px minimum, square crop)",
     PENDING),
    (15, "Upload 8–10 Work Photos",
     "Consultation session photos, before/after charts, vastu compass images, home diagrams",
     PENDING),
    (16, "Pre-populate Q&A (10 questions)",
     "Add these yourself: 'What is VIDS™?', 'Do you do online vastu?', 'What areas do you serve?', 'How much does a consultation cost?', 'Do you require demolition for corrections?'",
     PENDING),
    (17, "Enable Messaging",
     "Turn on GMB messaging — also mention WhatsApp +91 97391 05574 in posts",
     PENDING),
    (18, "First GMB Post",
     "Post: '5 Signs Your Home Needs a Vastu Check' — link to /vastu-for-bedroom or homepage",
     PENDING),
    (19, "Request first 5 reviews",
     "Send WhatsApp/email to past clients with direct GMB review link",
     PENDING),
    (20, "Respond to all reviews within 24h",
     "Ongoing — positive: thank + mention VIDS™. Negative: acknowledge + offer to resolve offline",
     PENDING),
]

for r, (num, task, what, status) in enumerate(gmb_tasks, 4):
    ws_gmb.row_dimensions[r].height = 40
    row_bg = LIGHT_GREEN if status == DONE else LIGHT_RED
    for col, val in enumerate([num, task, what, status, ""], 1):
        c = ws_gmb.cell(row=r, column=col, value=val)
        c.fill = fill(row_bg if col > 1 else GRAY)
        c.font = Font(color="111111", size=10, name="Calibri", bold=(col == 2))
        c.alignment = centre() if col in (1, 4) else left()
        c.border = border()
    dv_status.add(ws_gmb.cell(row=r, column=4))

ws_gmb.freeze_panes = "A4"

# ── Sheet 4: Rank Math Meta ──────────────────────────────────────────────────
ws_rm = wb.create_sheet("🔑 Rank Math Meta")
ws_rm.sheet_view.showGridLines = False
ws_rm.column_dimensions["A"].width = 38
ws_rm.column_dimensions["B"].width = 65
ws_rm.column_dimensions["C"].width = 55
ws_rm.column_dimensions["D"].width = 28
ws_rm.column_dimensions["E"].width = 18

ws_rm.merge_cells("A1:E1")
c = ws_rm["A1"]
c.value = "Rank Math Meta — SEO Title, Meta Description & Focus Keyword per Page"
c.fill = fill(NAVY); c.font = font(bold=True, size=14); c.alignment = centre()
ws_rm.row_dimensions[1].height = 40

ws_rm.merge_cells("A2:E2")
c = ws_rm["A2"]
c.value = "Set these in Rank Math for each page to achieve 90+ score. Status = ✅ once set."
c.fill = fill(GOLD); c.font = font(size=11); c.alignment = centre()
ws_rm.row_dimensions[2].height = 25

rm_headers = ["Slug / Page", "SEO Title (max 60 chars)", "Meta Description (max 155 chars)", "Focus Keyword", "Status"]
for col, h in enumerate(rm_headers, 1):
    c = ws_rm.cell(row=3, column=col, value=h)
    c.fill = fill(NAVY); c.font = font(bold=True); c.alignment = centre(); c.border = border()
ws_rm.row_dimensions[3].height = 28

rm_pages = [
    # Rooms
    ("/vastu-for-kitchen",         "Vastu for Kitchen: 7 Essential Rules for the Agni Zone",             "Vastu for kitchen — SE placement, cook facing East, stove position, sink placement, and 7 rules for a healthy, prosperous kitchen energy field.",                    "vastu for kitchen"),
    ("/vastu-for-bedroom",         "7 Essential Vastu for Bedroom Rules | Vardhini Vastu",                "Vastu for bedroom: SW zone, correct head direction, mirror rules & TV placement. Zero-demolition remedies by Raghavendra Hebbur, Bangalore.",                   "vastu for bedroom"),
    ("/vastu-for-living-room",     "7 Proven Vastu for Living Room Tips | Vardhini Vastu",                "Vastu for living room: zone placement, sofa direction, TV corner & clutter-free Brahmasthana. Zero-demolition remedies by Raghavendra Hebbur.",                  "vastu for living room"),
    ("/vastu-for-dining-room",     "Vastu for Dining Room: Direction, Table & Seating Rules | Vardhini",  "Vastu for dining room — best direction, dining table shape and placement, seating positions, and remedies for common dining room vastu defects.",               "vastu for dining room"),
    ("/vastu-for-bathroom",        "Vastu for Bathroom: Direction, Placement & Rules | Vardhini Vastu",   "Vastu for bathroom — NW placement, toilet direction, mirror rules & drain placement. Zero-demolition remedies for bathroom vastu defects.",                     "vastu for bathroom"),
    ("/vastu-for-pooja-room",      "Vastu for Pooja Room: Sacred Space Rules | Vardhini Vastu",           "Vastu for pooja room — NE zone, idol placement, lighting & sacred space rules for maximum divine grace energy in your home mandir.",                           "vastu for pooja room"),
    ("/vastu-for-main-entrance",   "Vastu for Main Entrance: Door Direction & Rules | Vardhini Vastu",    "Vastu for main entrance — best door direction, threshold rules, nameplate, and remedies for inauspicious entrance placements.",                                "vastu for main entrance"),
    ("/vastu-for-study-room",      "Vastu for Study Room: Direction, Desk & Focus Rules | Vardhini",      "Vastu for study room — NE or North zone, East-facing desk, bookshelf placement & rules for concentration, memory and academic success.",                       "vastu for study room"),
    ("/vastu-for-guest-room",      "Vastu for Guest Room: Direction & Placement | Vardhini Vastu",        "Vastu for guest room — NW zone, correct head direction and rules for welcoming guests without disturbing main family energy.",                                  "vastu for guest room"),
    ("/vastu-for-childrens-room",  "Vastu for Children's Room: Direction, Bed & Study Rules",             "Vastu for children's room — West zone, study desk direction, bed placement and colour rules for growth, focus and health.",                                    "vastu for children's room"),
    ("/vastu-for-master-bedroom",  "Vastu for Master Bedroom: SW Zone Rules | Vardhini Vastu",            "Vastu for master bedroom — SW zone placement, head direction, mirror & TV rules for the head of household's authority and deep rest.",                         "vastu for master bedroom"),
    ("/vastu-for-mandir-at-home",  "Vastu for Mandir at Home: Sacred Placement Rules | Vardhini",         "Vastu for mandir at home — NE zone, idol height, facing direction, light rules and remedies for in-home temple placement.",                                    "vastu for mandir at home"),
    ("/vastu-for-balcony",         "Vastu for Balcony: Direction & Placement Rules | Vardhini Vastu",     "Vastu for balcony — North or East balcony rules, plant placement, furniture direction and what to avoid in balcony vastu.",                                   "vastu for balcony"),
    ("/vastu-for-mirrors",         "Vastu for Mirrors: Placement & Rules | Vardhini Vastu",               "Vastu for mirrors — which walls are correct, which are defects, and the most dangerous mirror placements in bedroom, living room and dining room.",            "vastu for mirrors"),
    ("/vastu-for-shoe-rack",       "Vastu for Shoe Rack: Placement & Direction Rules | Vardhini Vastu",   "Vastu for shoe rack — correct zone, direction and placement rules to avoid negative energy entry through footwear storage.",                                  "vastu for shoe rack"),
    ("/vastu-for-staircase",       "Vastu for Staircase: Direction, Placement & Rules | Vardhini Vastu",  "Vastu for staircase — South or West zone, clockwise turn rule, under-staircase rules and corrections for inauspiciously placed stairs.",                      "vastu for staircase"),
    ("/vastu-for-basement",        "Vastu for Basement: Direction & Usage Rules | Vardhini Vastu",         "Vastu for basement — NW or North preferred, usage rules, lighting requirements and corrections for basements in inauspicious zones.",                         "vastu for basement"),
    ("/vastu-for-car-parking",     "Vastu for Car Parking: Direction & Placement Rules | Vardhini",        "Vastu for car parking — NW or SE zone, slope rules, gate direction and remedies for parking in inauspicious zones.",                                         "vastu for car parking"),
    ("/vastu-for-garden",          "Vastu for Garden: Direction, Plants & Layout Rules | Vardhini Vastu",  "Vastu for garden — North and East for open lawn, heavy trees in SW, water features in NE and plant selection for each zone.",                               "vastu for garden"),
    ("/vastu-for-water-tank",      "Vastu for Water Tank: Direction & Placement Rules | Vardhini Vastu",  "Vastu for water tank — NE for underground, SW for overhead, rules for size, shape and slope to maintain correct water element energy.",                     "vastu for water tank"),
    ("/vastu-for-swimming-pool",   "Vastu for Swimming Pool: Direction & Placement | Vardhini Vastu",     "Vastu for swimming pool — North or East zone, shape rules, and remedies for pools in inauspicious zones.",                                                   "vastu for swimming pool"),
    ("/vastu-for-septic-tank",     "Vastu for Septic Tank: Direction & Placement Rules | Vardhini Vastu", "Vastu for septic tank — NW or SW zone rules, distance from NE, and the most critical placement mistakes to avoid.",                                          "vastu for septic tank"),
    ("/vastu-for-wall-clock",      "Vastu for Wall Clock: Direction & Placement Rules | Vardhini Vastu",  "Vastu for wall clock — North wall primary, East acceptable, what to avoid and the significance of clock direction in vastu shastra.",                       "vastu for wall clock"),
    ("/vastu-for-colors",          "Vastu for Colors: Room-wise Colour Rules | Vardhini Vastu",            "Vastu for colors — zone-specific colour guide for bedroom, kitchen, living room, study room and entrance. Which colours to use and avoid in each direction.", "vastu for colors"),
    ("/vastu-for-toilet",          "Vastu for Toilet: Direction & Placement Rules | Vardhini Vastu",       "Vastu for toilet — NW or West zone, toilet seat direction, door rules and corrections for toilets in inauspicious zones like NE and SW.",                   "vastu for toilet"),
    # Key localities
    ("/vastu-consultant-whitefield",    "Vastu Consultant in Whitefield Bangalore | Vardhini Vastu",        "Vastu consultant in Whitefield — VIDS™ residential & commercial vastu by Raghavendra Hebbur. Zero-demolition corrections. Call +91 97391 05574.",  "vastu consultant in whitefield"),
    ("/vastu-consultant-indiranagar",   "Vastu Consultant in Indiranagar Bangalore | Vardhini Vastu",       "Vastu consultant in Indiranagar — VIDS™ system by Raghavendra Hebbur. Residential & commercial vastu. Zero-demolition corrections.",              "vastu consultant in indiranagar"),
    ("/vastu-consultant-koramangala",   "Vastu Consultant in Koramangala Bangalore | Vardhini Vastu",       "Vastu consultant in Koramangala — VIDS™ 16-zone analysis, zero-demolition corrections by Raghavendra Hebbur.",                                    "vastu consultant in koramangala"),
    # Cities
    ("/vastu-consultant-mumbai",        "Vastu Consultant in Mumbai | Vardhini Vastu — Online & In-Person", "Vastu consultant in Mumbai — VIDS™ system by Raghavendra Hebbur. Online vastu consultations + in-person visits. Call +91 97391 05574.",          "vastu consultant in mumbai"),
    ("/vastu-consultant-delhi",         "Vastu Consultant in Delhi | Vardhini Vastu — Online & In-Person",  "Vastu consultant in Delhi — degree-accurate VIDS™ analysis, zero-demolition remedies by Raghavendra Hebbur. Online + in-person.",               "vastu consultant in delhi"),
    ("/vastu-consultant-chennai",       "Vastu Consultant in Chennai | Vardhini Vastu",                     "Vastu consultant in Chennai — VIDS™ system by Raghavendra Hebbur. Online & in-person consultations. Zero-demolition corrections.",               "vastu consultant in chennai"),
]

for r, (slug, title, desc, kw) in enumerate(rm_pages, 4):
    ws_rm.row_dimensions[r].height = 40
    for col, val in enumerate([slug, title, desc, kw, PENDING], 1):
        c = ws_rm.cell(row=r, column=col, value=val)
        c.fill = fill(LIGHT_RED if col == 5 else (LIGHT_BLUE if r % 2 == 0 else WHITE))
        c.font = Font(color="111111", size=10, name="Calibri")
        c.alignment = left()
        c.border = border()
    dv_status.add(ws_rm.cell(row=r, column=5))

ws_rm.freeze_panes = "A4"

# ── Save ─────────────────────────────────────────────────────────────────────
out = r"C:/Users/raghu/VV/VardhiniVastu_SEO_Tracker.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total tasks: {len(tasks)} | Done: {done} | Pending: {pending}")
