# VardhiniVastu.in - SEO Master Tracker v2 - Full Strategy Build
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# Colours
NAVY="1E3A5F"; GOLD="B8860B"; GREEN="2D6A4F"; AMBER="D4740E"; RED="C53030"
CREAM="FAF9F7"; LT_BLUE="D6E4F0"; LT_GREEN="D5ECD8"; LT_RED="F9DEDE"
LT_AMBER="FAF0DC"; WHITE="FFFFFF"; GRAY="E8E8E8"

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False,color=WHITE,size=11): return Font(bold=bold,color=color,size=size,name="Calibri")
def bdr():
    t=Side(style="thin",color="CCCCCC"); return Border(left=t,right=t,top=t,bottom=t)
def ctr(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def lft(): return Alignment(horizontal="left",vertical="center",wrap_text=True)

DONE="Done"; PENDING="Pending"; INPROG="In Progress"

# ── Category colour map ──────────────────────────────────────────────────────
CAT_CLR = {
    "Strategy & Competitors":       ("1A3A6B","D6E4F7"),
    "Technical SEO":                ("1B4332","D5ECD8"),
    "Site Parity .com vs .in":      ("4A235A","EAD5F5"),
    "Footer & Templates":           ("5C3317","FDEBD0"),
    "Residential Rooms (25)":       ("6E2F0A","FAE5D3"),
    "Bangalore Core Localities":    ("145A32","D5F5E3"),
    "Bangalore Extended - South":   ("0E6655","D1F2EB"),
    "Bangalore Extended - East":    ("154360","D6EAF8"),
    "Bangalore Extended - North":   ("1B2631","D5D8DC"),
    "Bangalore Extended - West/NW": ("212F3D","D6DBDF"),
    "Apartment Targeting - Blore":  ("6C3483","F5EEF8"),
    "Apartment Targeting - Cities": ("1A5276","D6EAF8"),
    "Indian Cities (15)":           ("0B5345","D5F5E3"),
    "Industrial Areas - Bangalore": ("784212","FAE5D3"),
    "Industrial Areas - India":     ("6E2C00","FDEBD0"),
    "International (8)":            ("1B4F72","D6EAF8"),
    "Educational Hub":              ("4D1F00","FAE5D3"),
    "House Directions":             ("212F3D","D6DBDF"),
    "Core Pages":                   ("1C2833","D5D8DC"),
    "AI Citation & GEO":            ("0D47A1","E3F2FD"),
    "GMB & Local SEO":              ("C53030","FADBD8"),
    "Map Rank Tracker":             ("B7950B","FEF9E7"),
    "Google News & RSS":            ("1565C0","E3F2FD"),
    "Off-Page & Backlinks":         ("1A5276","D6EAF8"),
    "Analytics & Tracking":         ("117A65","D1F2EB"),
    "Daily Conversion Engine":      ("922B21","FADBD8"),
}

PRI_BG={"High":"FFDEDE","Medium":"FFF9DC","Low":"E2F5E2"}
STA_BG={DONE:LT_GREEN,PENDING:LT_RED,INPROG:LT_AMBER}

# ─────────────────────────────────────────────────────────────────────────────
# MASTER TASK DATA
# (Category, Task, Priority, Status, Notes)
# ─────────────────────────────────────────────────────────────────────────────
tasks = [

  # ── 1. Strategy & Competitors ────────────────────────────────────────────
  ("Strategy & Competitors","Identify top 5 Vastu consultants competing in Bangalore","High",DONE,"Competitors mapped: Mystic Space, VastuPlus, Vastu Devayah Namah, Astro Vastu, local listings"),
  ("Strategy & Competitors","Map competitor keyword universe — locality, city, room, direction pages","High",DONE,"Full roadmap benchmarked"),
  ("Strategy & Competitors","SEO content roadmap: Rooms + Localities + Cities + International + Edu + Directions","High",DONE,"68+ pages planned and executed"),
  ("Strategy & Competitors","Rank Math 90+ blueprint documented","High",DONE,"In memory/feedback_rankmath_blueprint.md"),
  ("Strategy & Competitors","Set Rank Math focus KW + SEO title + meta desc on all 68 live pages","High",PENDING,"Each page has metadata in HTML comment — paste into Rank Math for every page"),
  ("Strategy & Competitors","Monthly keyword rank tracking — top 30 target keywords","Medium",PENDING,"GSC + Rank Math rank tracker"),
  ("Strategy & Competitors","Competitor backlink gap analysis","Medium",PENDING,"Run /seo backlinks against top 3 competitors"),
  ("Strategy & Competitors","SEMrush project setup — track domain, backlinks, position changes","High",PENDING,"User started from SEMrush — ensure project is live and alerts set"),

  # ── 2. Technical SEO ─────────────────────────────────────────────────────
  ("Technical SEO","vv-blank.php custom template — created and verified","High",DONE,"Serves standalone HTML pages without Astra theme wrapper"),
  ("Technical SEO","WP permalink /%postname%/ — confirmed clean URLs on all pages","High",DONE,""),
  ("Technical SEO","XML sitemap — verify includes all 68+ new pages, submit to GSC","High",PENDING,"Rank Math auto-generates — check sitemap.xml coverage then submit"),
  ("Technical SEO","robots.txt — no important locality/room pages blocked","High",PENDING,"vardhinivastu.in/robots.txt audit"),
  ("Technical SEO","SSL/HTTPS — verify no mixed content on any page","High",PENDING,""),
  ("Technical SEO","Core Web Vitals (LCP, CLS, INP) — baseline via PageSpeed Insights","High",PENDING,"Run on homepage + 3 locality pages"),
  ("Technical SEO","Mobile rendering — all pages correct on mobile","High",PENDING,"Test standalone HTML pages vs fragment pages in Astra"),
  ("Technical SEO","Canonical tags — no duplicate content across city/locality variations","Medium",PENDING,"Check Rank Math canonical on all pages"),
  ("Technical SEO","Internal linking — rooms to localities, localities to cities, all to CTA","Medium",PENDING,"Build cross-link structure throughout site"),
  ("Technical SEO","Image upload — upload actual WebP images for all 68 page placeholder img tags","Medium",PENDING,"Pages reference /wp-content/uploads/ paths — images not yet uploaded"),
  ("Technical SEO","404 audit after indexing — fix any broken internal links","Low",PENDING,"Use GSC Coverage report or Screaming Frog"),
  ("Technical SEO","Schema markup validation — all JSON-LD in room and locality pages passes","High",PENDING,"Run schema validator on 5 sample pages"),
  ("Technical SEO","hreflang — add en-IN and en tags for Indian audience clarity","Low",PENDING,""),

  # ── 3. Site Parity — vardhinivastu.com vs vardhinivastu.in ───────────────
  ("Site Parity .com vs .in","Audit vardhinivastu.com — map all pages that exist there vs .in","High",PENDING,"List every page on .com and confirm .in equivalent exists"),
  ("Site Parity .com vs .in","Replicate all high-performing .com pages onto .in","High",PENDING,"The .com ranks well — port its best content to .in"),
  ("Site Parity .com vs .in","301 redirect strategy — decide if .in should eventually replace .com or coexist","High",PENDING,"Consider canonical cross-domain strategy or full migration"),
  ("Site Parity .com vs .in","Ensure homepage on .in has equal or better SEO treatment than .com","High",PENDING,"Title, meta, H1, schema, CTA all need to match or exceed .com"),
  ("Site Parity .com vs .in","Ensure contact/booking page on .in matches .com","High",PENDING,"Conversion page parity is critical for daily calls"),
  ("Site Parity .com vs .in","Ensure testimonials/reviews on .in match .com","Medium",PENDING,"Social proof is an E-E-A-T signal — both domains need it"),
  ("Site Parity .com vs .in","Cross-link .com and .in strategically to build combined authority","Medium",PENDING,"Link from .com blog posts to .in service pages and vice versa"),
  ("Site Parity .com vs .in","Verify GSC has both domains verified and tracked separately","High",PENDING,"Monitor both .in and .com in GSC separately"),

  # ── 4. Footer & Templates ────────────────────────────────────────────────
  ("Footer & Templates","Add vardhinivastu.com footer to ALL vardhinivastu.in pages — it ranked .com in 24hrs","High",PENDING,"Footer contains key internal links, NAP, and trust signals — replicate exactly on .in"),
  ("Footer & Templates","Create global footer include — one place to update, applies to all pages","High",PENDING,"Either a WP template part or a PHP include in vv-blank.php for standalone pages"),
  ("Footer & Templates","Footer content: NAP, service links, locality links, city links, social, schema","High",PENDING,"NAP: Vardhini Vastu | Bangalore | +91 97391 05574 | vardhinivastu.in"),
  ("Footer & Templates","Add footer to all 25 residential room pages","High",PENDING,"Fragment pages need footer injected or appended"),
  ("Footer & Templates","Add footer to all 20 Bangalore locality pages","High",PENDING,"Standalone HTML pages — footer needs to be in the HTML file"),
  ("Footer & Templates","Add footer to all 15 Indian city pages","High",PENDING,""),
  ("Footer & Templates","Add footer to all 8 international pages","High",PENDING,""),
  ("Footer & Templates","Update vv-blank.php to auto-inject footer on all standalone HTML pages","High",PENDING,"One template change covers all 43 standalone pages at once"),

  # ── 5. Residential Rooms (25 pages — all live) ───────────────────────────
  ("Residential Rooms (25)","vastu-for-kitchen (WP ID 283)","High",DONE,"vardhinivastu.in/vastu-for-kitchen"),
  ("Residential Rooms (25)","vastu-for-bedroom (WP ID 284)","High",DONE,"vardhinivastu.in/vastu-for-bedroom"),
  ("Residential Rooms (25)","vastu-for-living-room (WP ID 285)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-dining-room (WP ID 286)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-bathroom (WP ID 287)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-pooja-room (WP ID 288)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-main-entrance (WP ID 289)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-study-room (WP ID 290)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-guest-room (WP ID 291)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-childrens-room (WP ID 292)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-master-bedroom (WP ID 293)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-mandir-at-home (WP ID 294)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-balcony (WP ID 295)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-mirrors (WP ID 296)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-shoe-rack (WP ID 297)","Low",DONE,""),
  ("Residential Rooms (25)","vastu-for-staircase (WP ID 298)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-basement (WP ID 299)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-car-parking (WP ID 300)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-garden (WP ID 301)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-water-tank (WP ID 302)","Medium",DONE,""),
  ("Residential Rooms (25)","vastu-for-swimming-pool (WP ID 303)","Low",DONE,""),
  ("Residential Rooms (25)","vastu-for-septic-tank (WP ID 304)","Low",DONE,""),
  ("Residential Rooms (25)","vastu-for-wall-clock (WP ID 305)","Low",DONE,""),
  ("Residential Rooms (25)","vastu-for-colors (WP ID 306)","High",DONE,""),
  ("Residential Rooms (25)","vastu-for-toilet (WP ID 307)","Medium",DONE,""),

  # ── 6. Bangalore Core Localities (20 live) ───────────────────────────────
  ("Bangalore Core Localities","vastu-consultant-whitefield","High",DONE,"Live"),
  ("Bangalore Core Localities","vastu-consultant-indiranagar","High",DONE,"Live"),
  ("Bangalore Core Localities","vastu-consultant-koramangala","High",DONE,"Live"),
  ("Bangalore Core Localities","vastu-consultant-hsr-layout","High",DONE,"Live"),
  ("Bangalore Core Localities","vastu-consultant-electronic-city","High",DONE,"Live"),
  ("Bangalore Core Localities","vastu-consultant-jp-nagar","High",DONE,"Live"),
  ("Bangalore Core Localities","vastu-consultant-jayanagar","High",DONE,"Live"),
  ("Bangalore Core Localities","vastu-consultant-bannerghatta-road","Medium",DONE,"WP ID 267"),
  ("Bangalore Core Localities","vastu-consultant-sarjapur-road","Medium",DONE,"WP ID 268"),
  ("Bangalore Core Localities","vastu-consultant-marathahalli","Medium",DONE,"WP ID 269"),
  ("Bangalore Core Localities","vastu-consultant-yelahanka","Medium",DONE,"WP ID 270"),
  ("Bangalore Core Localities","vastu-consultant-hebbal","Medium",DONE,"WP ID 271"),
  ("Bangalore Core Localities","vastu-consultant-rajajinagar","Medium",DONE,"WP ID 272"),
  ("Bangalore Core Localities","vastu-consultant-malleswaram","Medium",DONE,"WP ID 273"),
  ("Bangalore Core Localities","vastu-consultant-basavanagudi","Medium",DONE,"WP ID 274"),
  ("Bangalore Core Localities","vastu-consultant-btm-layout","Medium",DONE,"WP ID 275"),
  ("Bangalore Core Localities","vastu-consultant-banaswadi","Low",DONE,"WP ID 276"),
  ("Bangalore Core Localities","vastu-consultant-bellandur","Medium",DONE,"WP ID 277"),
  ("Bangalore Core Localities","vastu-consultant-kr-puram","Low",DONE,"WP ID 278"),
  ("Bangalore Core Localities","vastu-consultant-nagarbhavi","Low",DONE,"WP ID 279"),

  # ── 7. Bangalore Extended — South / Kanakapura / Ramanagara ─────────────
  ("Bangalore Extended - South","vastu-consultant-banashankari","High",PENDING,"Dense residential — near Jayanagar, high apartment density"),
  ("Bangalore Extended - South","vastu-consultant-uttarahalli","Medium",PENDING,"Growing residential suburb south of Banashankari"),
  ("Bangalore Extended - South","vastu-consultant-kengeri","Medium",PENDING,"West Bangalore — affordable housing corridor"),
  ("Bangalore Extended - South","vastu-consultant-rajarajeshwari-nagar","High",PENDING,"RR Nagar — one of Bangalore's fastest growing suburbs"),
  ("Bangalore Extended - South","vastu-consultant-gottigere","Medium",PENDING,"South Bangalore — near Bannerghatta Road"),
  ("Bangalore Extended - South","vastu-consultant-hulimavu","Medium",PENDING,"Between BTM and Bannerghatta Road"),
  ("Bangalore Extended - South","vastu-consultant-kanakapura-road","High",PENDING,"Major growth corridor — luxury apartments, NICE Road"),
  ("Bangalore Extended - South","vastu-consultant-bidadi","Medium",PENDING,"Satellite town on Mysore Road — Toyota plant area, industrial + residential"),
  ("Bangalore Extended - South","vastu-consultant-ramanagara","Medium",PENDING,"District HQ — Silk City — growing real estate"),
  ("Bangalore Extended - South","vastu-consultant-mysore-road","High",PENDING,"Entire Mysore Road corridor — Kengeri to Bidadi"),
  ("Bangalore Extended - South","vastu-consultant-subramanyapura","Low",PENDING,"Residential pocket near Kanakapura Road"),
  ("Bangalore Extended - South","vastu-consultant-konanakunte","Low",PENDING,"Near NICE Road junction"),

  # ── 8. Bangalore Extended — East / Hoskote ──────────────────────────────
  ("Bangalore Extended - East","vastu-consultant-kadugodi","High",PENDING,"Between Whitefield and Hoskote — IT / apartment corridor"),
  ("Bangalore Extended - East","vastu-consultant-varthur","High",PENDING,"Next to Whitefield — rapid apartment development"),
  ("Bangalore Extended - East","vastu-consultant-hoskote","High",PENDING,"District town east of Bangalore — industrial + residential growth"),
  ("Bangalore Extended - East","vastu-consultant-old-madras-road","Medium",PENDING,"OMR corridor — KR Puram to Hoskote"),
  ("Bangalore Extended - East","vastu-consultant-itpl-road","High",PENDING,"ITPL / Hoodi — major IT zone, high apartment density"),
  ("Bangalore Extended - East","vastu-consultant-hoodi","Medium",PENDING,"Between Marathahalli and Whitefield"),
  ("Bangalore Extended - East","vastu-consultant-brookefield","Medium",PENDING,"Premium residential near Whitefield"),
  ("Bangalore Extended - East","vastu-consultant-kundalahalli","Low",PENDING,"Near ITPL"),
  ("Bangalore Extended - East","vastu-consultant-doddanekundi","Low",PENDING,"Outer Ring Road — Marathahalli to Whitefield"),

  # ── 9. Bangalore Extended — North / Devanahalli ─────────────────────────
  ("Bangalore Extended - North","vastu-consultant-thanisandra","High",PENDING,"Fast growing north Bangalore — Manyata Tech Park area"),
  ("Bangalore Extended - North","vastu-consultant-kogilu","Medium",PENDING,"North of Hebbal — new residential development"),
  ("Bangalore Extended - North","vastu-consultant-jakkur","Medium",PENDING,"Near Yelahanka — residential + weekend homes"),
  ("Bangalore Extended - North","vastu-consultant-doddaballapur-road","Medium",PENDING,"NH-648 corridor — industrial and residential mix"),
  ("Bangalore Extended - North","vastu-consultant-devanahalli","High",PENDING,"Near Kempegowda International Airport — luxury villas, ITIR zone"),
  ("Bangalore Extended - North","vastu-consultant-bagalur","Medium",PENDING,"Aerospace park zone — emerging residential"),
  ("Bangalore Extended - North","vastu-consultant-nandi-hills-road","Low",PENDING,"Weekend home corridor — villa plots"),
  ("Bangalore Extended - North","vastu-consultant-airport-road-bangalore","High",PENDING,"NH-44 corridor — premium residential + commercial"),
  ("Bangalore Extended - North","vastu-consultant-kempapura","Low",PENDING,"Near Hebbal — residential pocket"),

  # ── 10. Bangalore Extended — West / Tumkur Road ─────────────────────────
  ("Bangalore Extended - West/NW","vastu-consultant-yeshwanthpur","High",PENDING,"Major Tumkur Road junction — dense commercial + residential"),
  ("Bangalore Extended - West/NW","vastu-consultant-peenya","High",PENDING,"Largest industrial estate in South Asia — commercial vastu opportunity"),
  ("Bangalore Extended - West/NW","vastu-consultant-dasarahalli","Medium",PENDING,"Industrial + residential on Tumkur Road"),
  ("Bangalore Extended - West/NW","vastu-consultant-jalahalli","Medium",PENDING,"Air Force Station area — residential"),
  ("Bangalore Extended - West/NW","vastu-consultant-mathikere","Medium",PENDING,"West Bangalore residential"),
  ("Bangalore Extended - West/NW","vastu-consultant-tumkur-road","High",PENDING,"Full NH-48 corridor page — Yeshwanthpur to Nelamangala"),
  ("Bangalore Extended - West/NW","vastu-consultant-nelamangala","Medium",PENDING,"Logistics hub town — warehouse + residential"),
  ("Bangalore Extended - West/NW","vastu-consultant-magadi-road","Medium",PENDING,"West Bangalore corridor"),
  ("Bangalore Extended - West/NW","vastu-consultant-vijayanagar","High",PENDING,"Dense residential — one of Bangalore's oldest planned areas"),
  ("Bangalore Extended - West/NW","vastu-consultant-tumkur","Medium",PENDING,"Tumkur city — district HQ, growing real estate"),

  # ── 11. Apartment Targeting — Bangalore (no builder brand names) ─────────
  ("Apartment Targeting - Blore","vastu-for-high-rise-apartments-bangalore","High",PENDING,"Target buyers of 15+ floor apartments in Whitefield, Hebbal, Sarjapur — no builder names"),
  ("Apartment Targeting - Blore","vastu-for-2bhk-apartment-bangalore","High",PENDING,"Biggest search segment — 2BHK vastu rules specific to compact apartments"),
  ("Apartment Targeting - Blore","vastu-for-3bhk-apartment-bangalore","High",PENDING,"3BHK buyers — SW master bedroom in apartment context"),
  ("Apartment Targeting - Blore","vastu-for-studio-apartment-bangalore","Medium",PENDING,"Studio/1BHK — young IT professionals target audience"),
  ("Apartment Targeting - Blore","vastu-for-penthouse-bangalore","Medium",PENDING,"Premium segment — top floor vastu rules"),
  ("Apartment Targeting - Blore","vastu-for-apartment-whitefield","High",PENDING,"Whitefield apartment buyers — corridor-specific apartment vastu"),
  ("Apartment Targeting - Blore","vastu-for-apartment-sarjapur-road","High",PENDING,"Sarjapur Road — dense apartment corridor"),
  ("Apartment Targeting - Blore","vastu-for-apartment-hebbal","High",PENDING,"North Bangalore — luxury apartment zone near airport road"),
  ("Apartment Targeting - Blore","vastu-for-apartment-devanahalli","Medium",PENDING,"Airport zone — premium villa and apartment buyers"),
  ("Apartment Targeting - Blore","vastu-for-apartment-kanakapura-road","Medium",PENDING,"South Bangalore growth corridor"),
  ("Apartment Targeting - Blore","vastu-for-under-construction-flat","High",PENDING,"Pre-purchase vastu check — target people buying during construction"),
  ("Apartment Targeting - Blore","vastu-for-ready-to-move-flat","High",PENDING,"Ready possession — buyers who cannot modify structure"),
  ("Apartment Targeting - Blore","vastu-for-resale-apartment","Medium",PENDING,"Resale buyers — check existing defects before purchase"),
  ("Apartment Targeting - Blore","vastu-consultation-before-buying-flat","High",PENDING,"Transactional intent page — 'should I check vastu before buying'"),

  # ── 12. Apartment Targeting — Indian Cities ──────────────────────────────
  ("Apartment Targeting - Cities","vastu-for-apartments-mumbai","High",PENDING,"BKC, Worli, Powai, Thane — Category A apartment vastu"),
  ("Apartment Targeting - Cities","vastu-for-apartments-delhi-ncr","High",PENDING,"Gurgaon sectors, Noida, Greater Noida — high-rise vastu"),
  ("Apartment Targeting - Cities","vastu-for-apartments-hyderabad","High",PENDING,"Hitech City, Gachibowli, Kondapur, Banjara Hills"),
  ("Apartment Targeting - Cities","vastu-for-apartments-chennai","High",PENDING,"OMR, ECR, Velachery, Anna Nagar"),
  ("Apartment Targeting - Cities","vastu-for-apartments-pune","High",PENDING,"Hinjewadi, Kharadi, Wakad, Baner"),
  ("Apartment Targeting - Cities","vastu-for-apartments-ahmedabad","Medium",PENDING,"SG Highway corridor — premium apartments"),
  ("Apartment Targeting - Cities","vastu-for-apartments-kolkata","Medium",PENDING,"New Town Rajarhat, EM Bypass"),
  ("Apartment Targeting - Cities","vastu-for-2bhk-apartment-india","High",PENDING,"National-level 2BHK page — broadest reach"),
  ("Apartment Targeting - Cities","vastu-for-3bhk-apartment-india","High",PENDING,"National-level 3BHK page"),
  ("Apartment Targeting - Cities","vastu-for-apartment-purchase-checklist","High",PENDING,"Pre-purchase vastu checklist — high commercial intent, shareable"),

  # ── 13. Indian Cities (15 — all live) ────────────────────────────────────
  ("Indian Cities (15)","vastu-consultant-chennai (WP 255)","High",DONE,""),
  ("Indian Cities (15)","vastu-consultant-mumbai (WP 263)","High",DONE,""),
  ("Indian Cities (15)","vastu-consultant-delhi (WP 258)","High",DONE,""),
  ("Indian Cities (15)","vastu-consultant-hyderabad (WP 260)","High",DONE,""),
  ("Indian Cities (15)","vastu-consultant-pune (WP 265)","High",DONE,""),
  ("Indian Cities (15)","vastu-consultant-ahmedabad (WP 254)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-kolkata (WP 261)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-gurugram (WP 259)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-noida (WP 264)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-coimbatore (WP 256)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-mysuru (WP 262)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-mangaluru (WP 261b)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-kochi (WP 280)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-jaipur (WP 281)","Medium",DONE,""),
  ("Indian Cities (15)","vastu-consultant-chandigarh (WP 282)","Low",DONE,""),

  # ── 14. Industrial Areas — Bangalore ────────────────────────────────────
  ("Industrial Areas - Bangalore","vastu-consultant-peenya-industrial-area","High",PENDING,"Largest industrial estate in Asia — factory and warehouse vastu"),
  ("Industrial Areas - Bangalore","vastu-consultant-electronic-city-phase-2","High",PENDING,"Phase 2 industrial / IT expansion beyond Phase 1"),
  ("Industrial Areas - Bangalore","vastu-consultant-bommasandra-industrial-area","High",PENDING,"KIADB industrial area — manufacturing belt"),
  ("Industrial Areas - Bangalore","vastu-consultant-jigani-industrial-area","High",PENDING,"South Bangalore KIADB — pharma, garment, manufacturing"),
  ("Industrial Areas - Bangalore","vastu-consultant-dobbaspet","Medium",PENDING,"Tumkur Road industrial zone — warehousing, logistics"),
  ("Industrial Areas - Bangalore","vastu-consultant-bidadi-industrial-area","Medium",PENDING,"Toyota plant zone — industrial vastu"),
  ("Industrial Areas - Bangalore","vastu-consultant-hoskote-industrial-area","Medium",PENDING,"East Bangalore — KIADB Hoskote zone"),
  ("Industrial Areas - Bangalore","vastu-for-factory","High",PENDING,"Generic factory vastu page — targets all industrial searchers"),
  ("Industrial Areas - Bangalore","vastu-for-warehouse","High",PENDING,"Warehouse/logistics vastu — high commercial intent"),
  ("Industrial Areas - Bangalore","vastu-for-it-office-bangalore","High",PENDING,"IT park office vastu — targets Whitefield, Electronic City, Manyata"),

  # ── 15. Industrial Areas — India ─────────────────────────────────────────
  ("Industrial Areas - India","vastu-for-industrial-area","High",PENDING,"National-level industrial vastu hub page"),
  ("Industrial Areas - India","vastu-consultant-faridabad-industrial","Medium",PENDING,"NCR industrial belt"),
  ("Industrial Areas - India","vastu-consultant-noida-industrial-sector","Medium",PENDING,"Noida sectors 58-63 industrial"),
  ("Industrial Areas - India","vastu-consultant-pune-midc","High",PENDING,"Pune MIDC — Chakan, Pimpri-Chinchwad, Ranjangaon"),
  ("Industrial Areas - India","vastu-consultant-chennai-industrial-corridor","High",PENDING,"OMR IT, Sriperumbudur, Oragadam industrial zones"),
  ("Industrial Areas - India","vastu-consultant-hyderabad-pharma-city","High",PENDING,"Pharma City — Genome Valley — specialized industrial vastu"),
  ("Industrial Areas - India","vastu-consultant-surat-industrial","Medium",PENDING,"Diamond and textile industry hub"),
  ("Industrial Areas - India","vastu-consultant-rajkot-industrial","Low",PENDING,"Engineering goods industrial area"),
  ("Industrial Areas - India","vastu-consultant-coimbatore-industrial","Medium",PENDING,"Textile capital — Coimbatore industrial vastu"),
  ("Industrial Areas - India","vastu-for-commercial-office","High",PENDING,"Commercial office vastu — national reach"),
  ("Industrial Areas - India","vastu-for-showroom","High",PENDING,"Retail showroom vastu — high commercial intent"),
  ("Industrial Areas - India","vastu-for-restaurant","High",PENDING,"Restaurant vastu — massive search volume"),
  ("Industrial Areas - India","vastu-for-hotel","Medium",PENDING,"Hotel vastu — hospitality industry"),

  # ── 16. International (8 — all live) ─────────────────────────────────────
  ("International (8)","vastu-consultant-usa","High",DONE,"Live"),
  ("International (8)","vastu-consultant-uk","High",DONE,"Live"),
  ("International (8)","vastu-consultant-canada","High",DONE,"Live"),
  ("International (8)","vastu-consultant-australia","High",DONE,"Live"),
  ("International (8)","vastu-consultant-uae","High",DONE,"Live"),
  ("International (8)","vastu-consultant-singapore","Medium",DONE,"Live"),
  ("International (8)","vastu-consultant-new-zealand","Low",DONE,"Live"),
  ("International (8)","vastu-consultant-germany","Low",DONE,"Live"),

  # ── 17. Educational Hub (needs rewrite) ──────────────────────────────────
  ("Educational Hub","Rewrite 15 Edu Hub pages from old format to vv-page-hero — then push","Medium",PENDING,"Files in pages/educational-hub/ — content is good, template needs upgrade"),
  ("Educational Hub","vastu-shastra-history","Medium",PENDING,""),
  ("Educational Hub","vastu-shastra-principles","Medium",PENDING,""),
  ("Educational Hub","vastu-five-elements","Medium",PENDING,""),
  ("Educational Hub","vastu-16-zones","Medium",PENDING,""),
  ("Educational Hub","vastu-for-wealth","High",PENDING,"High volume — priority rewrite"),
  ("Educational Hub","vastu-for-health","High",PENDING,"High volume — priority rewrite"),
  ("Educational Hub","vastu-for-marriage","High",PENDING,"High volume"),
  ("Educational Hub","vastu-for-career","High",PENDING,""),
  ("Educational Hub","vastu-remedies","High",PENDING,""),
  ("Educational Hub","vastu-dosha","Medium",PENDING,""),
  ("Educational Hub","vastu-for-flat","Medium",PENDING,""),
  ("Educational Hub","vastu-for-office","Medium",PENDING,""),
  ("Educational Hub","vastu-for-plot","Medium",PENDING,""),
  ("Educational Hub","vastu-for-new-home","High",PENDING,""),
  ("Educational Hub","VIDS-system-explainer — build fresh, showcase unique methodology","High",PENDING,"Core differentiator page — VIDS is unique to Vardhini Vastu"),

  # ── 18. House Directions ─────────────────────────────────────────────────
  ("House Directions","Rewrite 10 House Direction pages to vv-page-hero then push to WP","Medium",PENDING,""),
  ("House Directions","vastu-for-north-facing-house","High",PENDING,"Very high search volume"),
  ("House Directions","vastu-for-south-facing-house","High",PENDING,"High volume — controversial = high engagement"),
  ("House Directions","vastu-for-east-facing-house","High",PENDING,"High volume"),
  ("House Directions","vastu-for-west-facing-house","Medium",PENDING,""),
  ("House Directions","vastu-for-northeast-facing","Medium",PENDING,""),
  ("House Directions","vastu-for-northwest-facing","Medium",PENDING,""),
  ("House Directions","vastu-for-southeast-facing","Medium",PENDING,""),
  ("House Directions","vastu-for-southwest-facing","Medium",PENDING,"High intent — SW is the most feared direction"),
  ("House Directions","vastu-for-corner-plot","Medium",PENDING,"Niche but very high buyer intent"),

  # ── 19. Core Pages ───────────────────────────────────────────────────────
  ("Core Pages","Homepage — full Rank Math 90+ treatment: title, meta, H1, schema, CTA","High",PENDING,"Most important page — drives all conversions"),
  ("Core Pages","About / Raghavendra Hebbur — E-E-A-T author page with credentials, photo, LinkedIn","High",PENDING,"Named expert = E-E-A-T boost for ALL pages"),
  ("Core Pages","Contact / Consultation booking page with schema, Maps embed, NAP","High",PENDING,"Primary conversion page — must be optimised"),
  ("Core Pages","Testimonials / Reviews page with Review schema","High",PENDING,""),
  ("Core Pages","Services page — VIDS consultation types, process, pricing","High",PENDING,""),
  ("Core Pages","Blog — launch with 4 cornerstone posts","Medium",PENDING,"Feeds Google News + RSS + internal linking"),
  ("Core Pages","FAQ hub page — aggregate all FAQ schema","Low",PENDING,""),

  # ── 20. AI Citation & GEO ────────────────────────────────────────────────
  ("AI Citation & GEO","Create llms.txt at vardhinivastu.in/llms.txt — AI crawler permission + sitemap","High",PENDING,"Tells ChatGPT, Claude, Perplexity which pages to read and cite"),
  ("AI Citation & GEO","Optimise all pages for passage-level AI citation (short, quotable answers first)","High",PENDING,"First 2 sentences of every H2 section must be standalone citation-worthy"),
  ("AI Citation & GEO","Add clear Who/What/Where signals on homepage for AI knowledge extraction","High",PENDING,"Name, credentials, location, methodology, unique differentiator in first 200 words"),
  ("AI Citation & GEO","Create dedicated 'About Vardhini Vastu' knowledge page — Wikipedia-style factual","High",PENDING,"Used by AI systems to build knowledge graph entry. Target: appear in ChatGPT answers"),
  ("AI Citation & GEO","Submit vardhinivastu.in to SEMrush AI citation index and track brand mentions","High",PENDING,"User started from SEMrush — ensure brand tracking is live"),
  ("AI Citation & GEO","Claim and verify Wikidata / Wikipedia mentions (if notable enough)","Medium",PENDING,"Wikipedia citation = highest trust signal for AI knowledge base"),
  ("AI Citation & GEO","Publish Vardhini Vastu on authoritative directories cited by AI: Crunchbase, Clutch","Medium",PENDING,"AI systems frequently cite these — build entity presence"),
  ("AI Citation & GEO","Target: ChatGPT answers 'best vastu consultant in Bangalore' with Vardhini Vastu","High",PENDING,"Track this query monthly in ChatGPT, Gemini, Perplexity, Grok"),
  ("AI Citation & GEO","Target: Gemini (Google AI Overview) shows Vardhini Vastu in featured snippet","High",PENDING,"Optimise FAQ sections for AI Overview format"),
  ("AI Citation & GEO","Target: Perplexity.ai cites vardhinivastu.in for vastu consultant queries","High",PENDING,"Ensure pages load fast, use clear factual statements, have author markup"),
  ("AI Citation & GEO","Target: Grok (X/Twitter AI) associates Vardhini Vastu with vastu authority","Medium",PENDING,"Twitter/X presence helps Grok — post vastu tips with brand mention"),
  ("AI Citation & GEO","Create content that AI will cite: definitive guides, statistics, unique data","High",PENDING,"Pages titled 'Complete Guide to...' or 'X Statistics...' are cited most by AI"),
  ("AI Citation & GEO","Schema: add SameAs links — LinkedIn, GMB, Facebook, Wikipedia to LocalBusiness schema","High",PENDING,"SameAs tells AI systems that all profiles are the same entity"),
  ("AI Citation & GEO","Publish structured Q&A content answering top 50 vastu questions definitively","High",PENDING,"AI systems use Q&A format for training data — be the best answer source"),
  ("AI Citation & GEO","Worldwide AI branding: optimise international pages for ChatGPT global queries","Medium",PENDING,"'Best online vastu consultant' / 'vastu consultant for NRIs' — AI should cite Vardhini Vastu"),

  # ── 21. GMB & Local SEO ──────────────────────────────────────────────────
  ("GMB & Local SEO","Claim and verify Google Business Profile","High",PENDING,"business.google.com — Priority #1"),
  ("GMB & Local SEO","GMB: Primary category = Vastu Consultant","High",PENDING,""),
  ("GMB & Local SEO","GMB: Complete 750-char description with VIDS, Raghavendra Hebbur, Bangalore, zero-demolition","High",PENDING,""),
  ("GMB & Local SEO","GMB: Service area — all 20 core + 30 extended Bangalore localities + top cities","High",PENDING,""),
  ("GMB & Local SEO","GMB: Upload 10+ photos — cover, logo, consultation, before/after","High",PENDING,""),
  ("GMB & Local SEO","GMB: Add all services with descriptions (Residential, Commercial, Online, VIDS)","High",PENDING,""),
  ("GMB & Local SEO","GMB: Q&A — pre-populate 10 questions","Medium",PENDING,""),
  ("GMB & Local SEO","GMB: Enable messaging + set auto-reply","Medium",PENDING,""),
  ("GMB & Local SEO","GMB: Weekly posts — tip, offer, or educational content","Medium",PENDING,"Ongoing — 1 post/week minimum"),
  ("GMB & Local SEO","GMB: Get first 10 reviews from past clients","High",PENDING,"Target: 10+ reviews, 4.8+ stars in 90 days"),
  ("GMB & Local SEO","GMB: Respond to ALL reviews within 24 hours","High",PENDING,"Ongoing"),
  ("GMB & Local SEO","GMB: Add appointment / booking link to contact page","High",PENDING,""),
  ("GMB & Local SEO","NAP consistency — identical across website, GMB, all citations","High",PENDING,"Vardhini Vastu | Bangalore | +91 97391 05574"),
  ("GMB & Local SEO","Citations Tier 1: Justdial, Sulekha, UrbanClap, IndiaMart, NoBroker","High",PENDING,""),
  ("GMB & Local SEO","Citations Tier 2: Housing.com, Magicbricks, 99acres vastu section","Medium",PENDING,""),
  ("GMB & Local SEO","LocalBusiness schema on homepage and contact page","High",PENDING,"Already in locality/room pages — add to core pages"),
  ("GMB & Local SEO","Bing Places listing — import from GMB","Low",PENDING,""),
  ("GMB & Local SEO","Apple Maps listing","Low",PENDING,"mapsconnect.apple.com"),

  # ── 22. Map Rank Tracker ─────────────────────────────────────────────────
  ("Map Rank Tracker","Set up geo-grid rank tracking for 'vastu consultant bangalore' in Google Maps","High",PENDING,"Use DataForSEO or BrightLocal geo-grid — track 7x7 grid across Bangalore"),
  ("Map Rank Tracker","Baseline map rank: record current position for top 10 target search terms","High",PENDING,"Before GMB optimisation — get baseline, then track weekly"),
  ("Map Rank Tracker","Target rank: Top 3 in Google Maps Local Pack for 'vastu consultant bangalore'","High",PENDING,"Primary goal — map pack appearance = direct calls"),
  ("Map Rank Tracker","Target rank: Appear in map pack for 20 Bangalore locality searches","High",PENDING,"e.g., 'vastu consultant whitefield' map pack"),
  ("Map Rank Tracker","Track map ranking for 'vastu consultant [city]' in all 15 Indian cities","Medium",PENDING,"Secondary cities — monthly tracking"),
  ("Map Rank Tracker","Optimise GMB reviews for map rank — reviews are #2 local ranking factor","High",PENDING,"More reviews + higher rating = higher map pack position"),
  ("Map Rank Tracker","GMB post frequency — weekly posts boost map rank algorithm","Medium",PENDING,""),
  ("Map Rank Tracker","Check competitors in map pack monthly — monitor their GMB activity","Medium",PENDING,""),

  # ── 23. Google News & RSS ────────────────────────────────────────────────
  ("Google News & RSS","Enable RSS feed — verify vardhinivastu.in/feed/ is live and valid","High",PENDING,"WordPress auto-generates RSS — verify it is accessible"),
  ("Google News & RSS","Submit RSS feed to Google via GSC News Sitemap","High",PENDING,"GSC > Sitemaps > Add news sitemap URL"),
  ("Google News & RSS","Apply for Google News publisher inclusion","High",PENDING,"news.google.com/publications — need consistent publishing schedule first"),
  ("Google News & RSS","Set up news sitemap: /news-sitemap.xml via Rank Math News Sitemap","High",PENDING,"Rank Math supports news sitemaps — enable and configure"),
  ("Google News & RSS","Publish blog content in news-eligible format — dateline, author, category","High",PENDING,"Google News requires byline, publication date, news category"),
  ("Google News & RSS","Content calendar: 2 news-eligible posts per week minimum","High",PENDING,"Topics: vastu for new construction, seasonal vastu tips, VIDS case studies"),
  ("Google News & RSS","Submit RSS to Feedly, Flipboard, and Indian news aggregators","Medium",PENDING,"Distribution amplifies each post beyond Google alone"),
  ("Google News & RSS","Pinterest RSS integration — vastu tips auto-post from RSS","Low",PENDING,""),
  ("Google News & RSS","Ensure blog posts include strong keywords in title for Google News ranking","High",PENDING,"Google News titles must be specific, factual, keyword-rich"),

  # ── 24. Off-Page & Backlinks ─────────────────────────────────────────────
  ("Off-Page & Backlinks","Guest article on housing.com / 99acres / Magicbricks blog","High",PENDING,"Target: 1 guest post/month, link back to vardhinivastu.in"),
  ("Off-Page & Backlinks","Get listed in Vastu / Astrology expert directories (Astroyogi, Ganeshaspeaks)","Medium",PENDING,""),
  ("Off-Page & Backlinks","LinkedIn — Raghavendra Hebbur profile optimised with all credentials","High",PENDING,"E-E-A-T author signal — links back to site"),
  ("Off-Page & Backlinks","YouTube channel — 2 vastu tip videos/month, embed on pages","Medium",PENDING,""),
  ("Off-Page & Backlinks","PR / media coverage — approach home decor and lifestyle journalists","Low",PENDING,"1 press mention per quarter"),
  ("Off-Page & Backlinks","Build backlinks from apartment community forums (Commonfloor, Apnacomplex)","High",PENDING,"Residents searching for vastu — link in helpful answers"),
  ("Off-Page & Backlinks","Quora / Reddit vastu answers with links back to relevant pages","High",PENDING,"Answer 'vastu consultant bangalore' questions — 2 per week"),

  # ── 25. Analytics & Tracking ─────────────────────────────────────────────
  ("Analytics & Tracking","Google Search Console — verify site, submit sitemap, monitor Coverage","High",PENDING,""),
  ("Analytics & Tracking","GA4 — verify tracking fires on ALL 68+ pages including standalone HTML","High",PENDING,"Standalone HTML pages need GA4 script added manually"),
  ("Analytics & Tracking","Rank Math Analytics — connect to GSC for keyword rank tracking in WP","High",PENDING,""),
  ("Analytics & Tracking","Conversion tracking — WhatsApp click, contact form, call click events in GA4","High",PENDING,"Primary KPI = daily calls"),
  ("Analytics & Tracking","Monthly SEO report — rankings, traffic, GMB views, calls, conversions","Medium",PENDING,"Baseline at 30 days post-GMB setup"),
  ("Analytics & Tracking","Set up Google Alerts for 'Vardhini Vastu' and 'Raghavendra Hebbur'","Medium",PENDING,"Brand monitoring — see who is mentioning you"),
  ("Analytics & Tracking","SEMrush weekly alerts — keyword changes, backlink gains/losses","High",PENDING,""),

  # ── 26. Daily Conversion Engine ─────────────────────────────────────────
  ("Daily Conversion Engine","WhatsApp Business — set up auto-reply with consultation booking info","High",PENDING,"Every page has +91 97391 05574 — ensure WhatsApp Business is configured"),
  ("Daily Conversion Engine","Click-to-call tracking — ensure all phone numbers are tap-to-call on mobile","High",PENDING,"tel: links on all pages"),
  ("Daily Conversion Engine","Response time target: reply to all inquiries within 2 hours","High",PENDING,"Speed of response = conversion rate"),
  ("Daily Conversion Engine","Calendly or booking link on every page — reduce friction","High",PENDING,"Every page CTA should link to a direct booking page"),
  ("Daily Conversion Engine","GMB: Missed call auto-text — enable if available in your region","Medium",PENDING,""),
  ("Daily Conversion Engine","Retargeting pixel (Google/Meta) — track visitors, run retargeting ads","Medium",PENDING,"Visitors who don't convert get retargeted with consultation offer"),
  ("Daily Conversion Engine","Offer: Free 15-min vastu Q&A call — funnel to paid consultation","High",PENDING,"Low-friction entry offer increases contact rate"),
  ("Daily Conversion Engine","Review funnel: after every consultation, send WhatsApp review request","High",PENDING,"Systematise reviews — every happy client = 1 GMB review"),
  ("Daily Conversion Engine","Daily check: GSC clicks, GA4 sessions, WhatsApp messages, GMB calls","High",PENDING,"30-min daily review of key metrics to stay on top of momentum"),
]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_d = wb.active
ws_d.title = "Dashboard"
ws_d.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [32,20,20,20,35]):
    ws_d.column_dimensions[col].width = w

ws_d.row_dimensions[1].height = 60
ws_d.merge_cells("A1:E1")
c = ws_d["A1"]
c.value = "Vardhini Vastu  —  SEO Master Tracker  v2"
c.fill = fill(NAVY); c.font = Font(bold=True,color=WHITE,size=20,name="Calibri"); c.alignment = ctr()

ws_d.merge_cells("A2:E2")
c = ws_d["A2"]
c.value = f"vardhinivastu.in  &  vardhinivastu.com  |  Raghavendra Hebbur  |  Updated: {date.today().strftime('%d %b %Y')}"
c.fill = fill(GOLD); c.font = Font(color=WHITE,size=12,name="Calibri"); c.alignment = ctr()

total = len(tasks)
done_n   = sum(1 for t in tasks if t[3]==DONE)
pend_n   = sum(1 for t in tasks if t[3]==PENDING)
inp_n    = sum(1 for t in tasks if t[3]==INPROG)
pct      = int(done_n/total*100)

stats = [
    ("Total Tasks",            total,   "100%",   "Across all 26 categories"),
    ("Done",                   done_n,  f"{pct}%", "Live on site / process complete"),
    ("Pending",                pend_n,  f"{100-pct}%","Not yet started — prioritise by colour"),
    ("In Progress",            inp_n,   "—",       "Currently active"),
    ("Pages Live on WP",       68,      "—",       "Rooms 25 + Localities 20 + Cities 15 + International 8"),
    ("Bangalore Localities Total", 50,  "—",       "20 live + 30 extended (South/East/North/West corridors)"),
    ("Apartment Pages Pipeline",   24,  "—",       "Bangalore 14 + Cities 10 — no builder names"),
    ("Industrial Area Pages",      23,  "—",       "Bangalore 10 + India 13 — commercial vastu segment"),
    ("GMB Tasks (Highest Priority)",18, "—",       "Zero started — #1 lever for daily calls"),
    ("AI Citation Tasks",          15,  "—",       "ChatGPT, Gemini, Grok, Perplexity, SEMrush"),
]

ws_d.row_dimensions[4].height=12
for col in "ABCDE": ws_d.cell(4,ord(col)-64).fill = fill(CREAM)

hdr = ["Metric","Count","% Total","Notes"]
for ci,h in enumerate(hdr,1):
    c = ws_d.cell(5,ci,h); c.fill=fill(NAVY); c.font=font(True); c.alignment=ctr(); c.border=bdr()
    ws_d.row_dimensions[5].height=24

for ri,(label,cnt,pct_s,note) in enumerate(stats,6):
    ws_d.row_dimensions[ri].height = 22
    bg = LT_GREEN if "Done" in label else LT_RED if "Pending" in label else LT_BLUE
    for ci,val in enumerate([label,cnt,pct_s,note],1):
        c = ws_d.cell(ri,ci,val)
        c.fill = fill(bg); c.border = bdr()
        c.font = Font(bold=(ci==1),color="111111",size=11,name="Calibri")
        c.alignment = ctr() if ci in (2,3) else lft()

# Priority guide
ws_d.row_dimensions[17].height=12
for ci in range(1,6): ws_d.cell(17,ci).fill=fill(CREAM)

ws_d.merge_cells("A18:E18")
c=ws_d["A18"]; c.value="TOP PRIORITY ACTIONS (start here)"; c.fill=fill(RED)
c.font=Font(bold=True,color=WHITE,size=13,name="Calibri"); c.alignment=ctr()
ws_d.row_dimensions[18].height=30

priorities=[
    ("1","Claim & complete Google Business Profile","GMB & Local SEO","Direct calls = revenue"),
    ("2","Add vardhinivastu.com footer to all .in pages","Footer & Templates","Ranked .com in 24hrs"),
    ("3","Set Rank Math meta on all 68 live pages","Strategy & Competitors","Quick SEO wins"),
    ("4","Create llms.txt — AI citation foundation","AI Citation & GEO","ChatGPT/Gemini/Grok"),
    ("5","Submit XML sitemap + verify GSC","Technical SEO","Index all new pages"),
    ("6","Build 30 extended Bangalore locality pages","Bangalore Extended","Dominate micro-searches"),
    ("7","Apartment targeting pages (2BHK/3BHK etc)","Apartment Targeting - Blore","High intent buyers"),
    ("8","Industrial area pages — factory/warehouse vastu","Industrial Areas - Bangalore","Commercial revenue"),
    ("9","Weekly GMB posts + review collection","GMB & Local SEO","Map pack ranking"),
    ("10","Google News setup + 2 posts/week","Google News & RSS","First page news results"),
]
for ci,h in enumerate(["#","Action","Category","Why It Matters"],1):
    c=ws_d.cell(19,ci,h); c.fill=fill(NAVY); c.font=font(True); c.alignment=ctr(); c.border=bdr()
ws_d.row_dimensions[19].height=24

for ri,(n,action,cat,why) in enumerate(priorities,20):
    ws_d.row_dimensions[ri].height=22
    for ci,val in enumerate([n,action,cat,why],1):
        c=ws_d.cell(ri,ci,val)
        c.fill=fill("FFF3CD" if int(n)<=5 else LT_BLUE)
        c.font=Font(bold=(ci in(1,2)),color="111111",size=10,name="Calibri")
        c.alignment=ctr() if ci==1 else lft(); c.border=bdr()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: MASTER TASK LIST
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Master Task List")
ws.sheet_view.showGridLines = False

col_ws = [5,30,52,10,16,52,14,14]
for i,w in enumerate(col_ws,1): ws.column_dimensions[get_column_letter(i)].width=w

# Freeze + header
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 40
hdrs = ["#","Category","Task / Page","Priority","Status","Notes / URL","Target Date","Done Date"]
for ci,h in enumerate(hdrs,1):
    c=ws.cell(1,ci,h); c.fill=fill(NAVY); c.font=font(True); c.alignment=ctr(); c.border=bdr()

# Dropdowns
dv_s=DataValidation(type="list",formula1='"Done,In Progress,Pending"',allow_blank=False)
dv_p=DataValidation(type="list",formula1='"High,Medium,Low"',allow_blank=False)
ws.add_data_validation(dv_s); ws.add_data_validation(dv_p)

row=2; prev_cat=None; idx=0
for (cat,task,pri,sta,notes) in tasks:
    idx+=1
    if cat!=prev_cat:
        if prev_cat is not None:
            ws.row_dimensions[row].height=6
            for ci in range(1,9): ws.cell(row,ci).fill=fill(CREAM)
            row+=1
        ws.row_dimensions[row].height=26
        ws.merge_cells(f"A{row}:H{row}")
        hc,_=CAT_CLR.get(cat,(NAVY,WHITE))
        c=ws.cell(row,1,f"  {cat}")
        c.fill=fill(hc); c.font=Font(bold=True,color=WHITE,size=11,name="Calibri")
        c.alignment=Alignment(horizontal="left",vertical="center")
        row+=1; prev_cat=cat

    ws.row_dimensions[row].height=36
    _,bg=CAT_CLR.get(cat,(NAVY,"F5F5F5"))
    vals=[idx,cat,task,pri,sta,notes,"",""]
    for ci,val in enumerate(vals,1):
        c=ws.cell(row,ci,val)
        if ci==4: c.fill=fill(PRI_BG.get(pri,WHITE))
        elif ci==5: c.fill=fill(STA_BG.get(sta,WHITE))
        else: c.fill=fill(bg)
        c.font=Font(bold=(ci in(2,3)),color="111111",size=10,name="Calibri")
        c.alignment=ctr() if ci in(1,4,5) else lft()
        c.border=bdr()
    dv_s.add(ws.cell(row,5)); dv_p.add(ws.cell(row,4))
    row+=1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: BANGALORE LOCALITY PIPELINE
# ══════════════════════════════════════════════════════════════════════════════
ws_b = wb.create_sheet("Bangalore Locality Pipeline")
ws_b.sheet_view.showGridLines = False
for col,w in zip("ABCDE",[5,38,20,20,40]):
    ws_b.column_dimensions[col].width=w

ws_b.merge_cells("A1:E1")
c=ws_b["A1"]; c.value="Bangalore Locality Pipeline — 50 Localities Total"
c.fill=fill(NAVY); c.font=font(True,size=14); c.alignment=ctr()
ws_b.row_dimensions[1].height=40

ws_b.merge_cells("A2:E2")
c=ws_b["A2"]; c.value="20 Live + 30 Extended (South/Kanakapura | East/Hoskote | North/Devanahalli | West/Tumkur Road)"
c.fill=fill(GOLD); c.font=font(size=11); c.alignment=ctr()
ws_b.row_dimensions[2].height=24

for ci,h in enumerate(["#","Slug","Corridor","Status","Notes"],1):
    c=ws_b.cell(3,ci,h); c.fill=fill(NAVY); c.font=font(True); c.alignment=ctr(); c.border=bdr()
ws_b.row_dimensions[3].height=26

localities=[
  # Core (20 live)
  (1,"vastu-consultant-whitefield","Core — East",DONE,"WP Live"),
  (2,"vastu-consultant-indiranagar","Core — Central",DONE,"WP Live"),
  (3,"vastu-consultant-koramangala","Core — South",DONE,"WP Live"),
  (4,"vastu-consultant-hsr-layout","Core — South",DONE,"WP Live"),
  (5,"vastu-consultant-electronic-city","Core — South",DONE,"WP Live"),
  (6,"vastu-consultant-jp-nagar","Core — South",DONE,"WP Live"),
  (7,"vastu-consultant-jayanagar","Core — South",DONE,"WP Live"),
  (8,"vastu-consultant-bannerghatta-road","Core — South",DONE,"WP ID 267"),
  (9,"vastu-consultant-sarjapur-road","Core — East",DONE,"WP ID 268"),
  (10,"vastu-consultant-marathahalli","Core — East",DONE,"WP ID 269"),
  (11,"vastu-consultant-yelahanka","Core — North",DONE,"WP ID 270"),
  (12,"vastu-consultant-hebbal","Core — North",DONE,"WP ID 271"),
  (13,"vastu-consultant-rajajinagar","Core — West",DONE,"WP ID 272"),
  (14,"vastu-consultant-malleswaram","Core — West",DONE,"WP ID 273"),
  (15,"vastu-consultant-basavanagudi","Core — South",DONE,"WP ID 274"),
  (16,"vastu-consultant-btm-layout","Core — South",DONE,"WP ID 275"),
  (17,"vastu-consultant-banaswadi","Core — East",DONE,"WP ID 276"),
  (18,"vastu-consultant-bellandur","Core — East",DONE,"WP ID 277"),
  (19,"vastu-consultant-kr-puram","Core — East",DONE,"WP ID 278"),
  (20,"vastu-consultant-nagarbhavi","Core — West",DONE,"WP ID 279"),
  # Extended South (12)
  (21,"vastu-consultant-banashankari","South / Kanakapura",PENDING,"Dense residential"),
  (22,"vastu-consultant-uttarahalli","South / Kanakapura",PENDING,""),
  (23,"vastu-consultant-kengeri","South / Mysore Road",PENDING,""),
  (24,"vastu-consultant-rajarajeshwari-nagar","South / Mysore Road",PENDING,"RR Nagar — fast growth"),
  (25,"vastu-consultant-gottigere","South / Bannerghatta Rd",PENDING,""),
  (26,"vastu-consultant-hulimavu","South / Bannerghatta Rd",PENDING,""),
  (27,"vastu-consultant-kanakapura-road","South / Kanakapura",PENDING,"NICE Road corridor"),
  (28,"vastu-consultant-subramanyapura","South / Kanakapura",PENDING,""),
  (29,"vastu-consultant-konanakunte","South / Kanakapura",PENDING,""),
  (30,"vastu-consultant-mysore-road","South / Mysore Road",PENDING,"Full corridor page"),
  (31,"vastu-consultant-bidadi","South / Mysore Road",PENDING,"Satellite town — Toyota"),
  (32,"vastu-consultant-ramanagara","South / Mysore Road",PENDING,"District HQ"),
  # Extended East (9)
  (33,"vastu-consultant-kadugodi","East / Hoskote",PENDING,"IT corridor"),
  (34,"vastu-consultant-varthur","East / Whitefield adj",PENDING,""),
  (35,"vastu-consultant-hoskote","East / Hoskote",PENDING,"District town"),
  (36,"vastu-consultant-old-madras-road","East / Hoskote",PENDING,"OMR corridor"),
  (37,"vastu-consultant-itpl-road","East / ITPL",PENDING,"Major IT zone"),
  (38,"vastu-consultant-hoodi","East / Marathahalli adj",PENDING,""),
  (39,"vastu-consultant-brookefield","East / Whitefield adj",PENDING,"Premium residential"),
  (40,"vastu-consultant-kundalahalli","East / ITPL",PENDING,""),
  (41,"vastu-consultant-doddanekundi","East / ORR",PENDING,""),
  # Extended North (9)
  (42,"vastu-consultant-thanisandra","North / Manyata",PENDING,"Fast growing"),
  (43,"vastu-consultant-kogilu","North / Devanahalli",PENDING,""),
  (44,"vastu-consultant-jakkur","North / Yelahanka adj",PENDING,""),
  (45,"vastu-consultant-doddaballapur-road","North / NH648",PENDING,""),
  (46,"vastu-consultant-devanahalli","North / Airport",PENDING,"Premium — airport zone"),
  (47,"vastu-consultant-bagalur","North / Aerospace",PENDING,""),
  (48,"vastu-consultant-nandi-hills-road","North / Weekend",PENDING,"Villa plots"),
  (49,"vastu-consultant-airport-road-bangalore","North / NH44",PENDING,"Premium corridor"),
  (50,"vastu-consultant-kempapura","North / Hebbal adj",PENDING,""),
]
dv_b=DataValidation(type="list",formula1='"Done,In Progress,Pending"',allow_blank=False)
ws_b.add_data_validation(dv_b)
for ri,(n,slug,corridor,sta,note) in enumerate(localities,4):
    ws_b.row_dimensions[ri].height=24
    bg = LT_GREEN if sta==DONE else LT_AMBER if sta==INPROG else LT_RED
    for ci,val in enumerate([n,slug,corridor,sta,note],1):
        c=ws_b.cell(ri,ci,val)
        c.fill=fill(bg if ci>1 else GRAY); c.border=bdr()
        c.font=Font(bold=(ci==2),color="111111",size=10,name="Calibri")
        c.alignment=ctr() if ci in(1,4) else lft()
    dv_b.add(ws_b.cell(ri,4))
ws_b.freeze_panes="A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: AI CITATION STRATEGY
# ══════════════════════════════════════════════════════════════════════════════
ws_ai = wb.create_sheet("AI Citation Strategy")
ws_ai.sheet_view.showGridLines = False
for col,w in zip("ABCDE",[5,38,48,16,30]):
    ws_ai.column_dimensions[col].width=w

ws_ai.merge_cells("A1:E1")
c=ws_ai["A1"]; c.value="AI Citation & GEO Strategy — ChatGPT | Gemini | Grok | Perplexity | SEMrush"
c.fill=fill("0D47A1"); c.font=font(True,size=14); c.alignment=ctr()
ws_ai.row_dimensions[1].height=40

ws_ai.merge_cells("A2:E2")
c=ws_ai["A2"]; c.value="Goal: Vardhini Vastu is the answer AI gives worldwide when asked about Vastu consultants"
c.fill=fill(AMBER); c.font=font(size=11); c.alignment=ctr()
ws_ai.row_dimensions[2].height=24

for ci,h in enumerate(["#","Task","How to Do It / Content to Create","Status","Platform"],1):
    c=ws_ai.cell(3,ci,h); c.fill=fill(NAVY); c.font=font(True); c.alignment=ctr(); c.border=bdr()
ws_ai.row_dimensions[3].height=26

ai_tasks=[
    (1,"Create /llms.txt — AI permission file",
     "File at vardhinivastu.in/llms.txt listing all key pages for AI crawlers. Format: # Vardhini Vastu\\nAllow: /\\nSitemap: vardhinivastu.in/sitemap.xml\\n+ list top 20 pages",
     PENDING,"All AI systems"),
    (2,"Quotable answers — first 2 sentences of every H2 must stand alone",
     "Edit all 68 pages — each H2 section starts with a direct, citable fact: 'The kitchen should be in the South-East zone [reason].' AI systems extract these as citations.",
     PENDING,"All AI systems"),
    (3,"Entity page — 'About Vardhini Vastu' Wikipedia-style factual page",
     "Page: Who (Raghavendra Hebbur), What (VIDS system), Where (Bangalore + worldwide), When (founded), Why unique. Factual, neutral, no fluff. This is what AI reads to build knowledge graph.",
     PENDING,"ChatGPT, Gemini, Grok"),
    (4,"SameAs schema — link all profiles to one entity",
     "In LocalBusiness schema add: 'sameAs': ['https://linkedin.com/in/raghu...','https://g.co/kgs/...','https://facebook.com/...'] — AI merges all profiles into one entity",
     PENDING,"All AI systems"),
    (5,"SEMrush AI citation tracking — set up brand monitoring",
     "SEMrush Brand Monitoring tool: track 'Vardhini Vastu' + 'Raghavendra Hebbur' mentions across web. User already started from SEMrush — activate this module.",
     PENDING,"SEMrush"),
    (6,"Wikidata entry for Vardhini Vastu",
     "Create Wikidata item: Vardhini Vastu | instance of: business | country: India | website: vardhinivastu.in | topic: Vastu Shastra. Wikidata = primary AI knowledge source.",
     PENDING,"All AI — Wikidata"),
    (7,"Clutch.co + Crunchbase listing",
     "AI systems cite Clutch for service businesses. Create detailed profile: services, team size, reviews, speciality. Crunchbase: org profile with description.",
     PENDING,"ChatGPT, Perplexity"),
    (8,"Publish 'Top 50 Vastu Questions' definitive guide",
     "Page targeting: 'vastu for bedroom', 'vastu for kitchen' etc — 50 Q&As, each with a clear 2-sentence answer. AI systems use Q&A format — this page will be cited heavily.",
     PENDING,"All AI systems"),
    (9,"Gemini / AI Overview optimisation",
     "For each room/locality page: add a clear 40-word definition at top of page. Google AI Overview pulls from the clearest, most concise answer. Target: featured in AI Overview for 'vastu for kitchen' etc.",
     PENDING,"Google Gemini"),
    (10,"ChatGPT Plugin / GPT Store presence",
     "Create a GPT or prompt that references Vardhini Vastu. Or ensure vardhinivastu.in is cited in ChatGPT's web browsing results for vastu queries.",
     PENDING,"ChatGPT"),
    (11,"Grok (X/Twitter AI) — build Twitter/X presence",
     "Post vastu tips 3x/week on X with #vastu #vastutiips #vastuexpert. Grok trains on X data — consistent brand presence on X = Grok brand recognition.",
     PENDING,"Grok / X AI"),
    (12,"Perplexity.ai — ensure pages load fast and have clear citations",
     "Perplexity prefers pages with: fast load (<2s), clear author, date, source links, and direct answers. Run PageSpeed on key pages and add author + date markup.",
     PENDING,"Perplexity"),
    (13,"Worldwide AI visibility — online vastu consultant query",
     "Target global ChatGPT/Gemini answers: 'best online vastu consultant India' / 'vastu consultant for NRIs'. Build international pages with clear online consultation offering.",
     PENDING,"All AI — Global"),
    (14,"Press mentions and news articles about Vardhini Vastu",
     "AI systems weight news mentions heavily. Target: 1 media mention per quarter in a real estate or lifestyle publication. Even a quote in a blog counts.",
     PENDING,"All AI systems"),
    (15,"Consistent NAP + VIDS terminology in all content",
     "AI builds brand knowledge from repeated consistent signals. Every page should mention: Vardhini Vastu | Raghavendra Hebbur | VIDS system | Bangalore | zero-demolition. Consistency = entity strength.",
     PENDING,"All AI systems"),
]
dv_ai=DataValidation(type="list",formula1='"Done,In Progress,Pending"',allow_blank=False)
ws_ai.add_data_validation(dv_ai)
for ri,(n,task,how,sta,platform) in enumerate(ai_tasks,4):
    ws_ai.row_dimensions[ri].height=52
    bg=LT_GREEN if sta==DONE else LT_RED
    for ci,val in enumerate([n,task,how,sta,platform],1):
        c=ws_ai.cell(ri,ci,val)
        c.fill=fill(bg if ci>1 else GRAY); c.border=bdr()
        c.font=Font(bold=(ci==2),color="111111",size=10,name="Calibri")
        c.alignment=ctr() if ci in(1,4) else lft()
    dv_ai.add(ws_ai.cell(ri,4))
ws_ai.freeze_panes="A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: GMB QUICK-START
# ══════════════════════════════════════════════════════════════════════════════
ws_g = wb.create_sheet("GMB Quick-Start")
ws_g.sheet_view.showGridLines = False
for col,w in zip("ABCDE",[5,38,52,16,16]):
    ws_g.column_dimensions[col].width=w

ws_g.merge_cells("A1:E1")
c=ws_g["A1"]; c.value="Google Business Profile — Quick-Start Checklist (Complete in Order)"
c.fill=fill(RED); c.font=font(True,size=14); c.alignment=ctr()
ws_g.row_dimensions[1].height=40

ws_g.merge_cells("A2:E2")
c=ws_g["A2"]; c.value="GMB = #1 local ranking factor for daily calls. Complete ALL steps before touching anything else."
c.fill=fill(AMBER); c.font=font(size=11); c.alignment=ctr()
ws_g.row_dimensions[2].height=24

for ci,h in enumerate(["#","Task","What to Write / Do","Status","Done Date"],1):
    c=ws_g.cell(3,ci,h); c.fill=fill(NAVY); c.font=font(True); c.alignment=ctr(); c.border=bdr()
ws_g.row_dimensions[3].height=26

gmb=[
    (1,"Claim & verify GMB listing","business.google.com → Add business → Verify by postcard or phone call",PENDING),
    (2,"Business Name","Vardhini Vastu  (no extra keywords — Google penalises keyword-stuffed names)",PENDING),
    (3,"Primary Category","Vastu Consultant",PENDING),
    (4,"Additional Categories","Feng Shui Consultant | Spiritual Counselor | Home Consultant",PENDING),
    (5,"Business Description (750 chars)","Vardhini Vastu offers professional Vastu Shastra consultancy by Raghavendra Hebbur in Bangalore. Specialising in the VIDS system (Vardhini Integrated Direction System) — a 16-zone degree-accurate Vastu analysis with zero-demolition corrections. Services include residential vastu, commercial vastu, plot analysis, and online consultations across India and worldwide.",PENDING),
    (6,"Phone Number","+91 97391 05574  (must match website and all citations exactly — NAP consistency)",PENDING),
    (7,"Website URL","https://vardhinivastu.in",PENDING),
    (8,"Address / Service Area","Full Bangalore address OR hide address + set service area. Add all 50 localities.",PENDING),
    (9,"Opening Hours","Mon-Sat 10am-7pm | Sun By Appointment  (accurate hours = higher call rate)",PENDING),
    (10,"Services — add each","1) Residential Vastu 2) Commercial Vastu 3) Online Consultation 4) Plot/Land Vastu 5) VIDS Analysis",PENDING),
    (11,"Booking / Appointment Link","https://vardhinivastu.in/vastu-consultant-bangalore-contact/",PENDING),
    (12,"Cover Photo","Professional headshot of Raghavendra Hebbur or clean vastu compass image (1332x750px)",PENDING),
    (13,"Logo","Vardhini Vastu logo (250x250px minimum, square)",PENDING),
    (14,"Work Photos (8-10)","Consultation session, vastu charts, before/after home layouts, compass analysis",PENDING),
    (15,"Q&A — pre-populate 10 questions","Add: What is VIDS? / Do you do online vastu? / Areas served? / Cost? / No demolition?",PENDING),
    (16,"Enable Messaging","Turn on GMB messaging — mention WhatsApp +91 97391 05574 in auto-reply",PENDING),
    (17,"First GMB Post","Post: '5 Signs Your Home Needs a Vastu Check' — link to /vastu-for-bedroom",PENDING),
    (18,"Request first 5-10 reviews","WhatsApp past clients: 'Could you leave a quick Google review? [GMB link]'",PENDING),
    (19,"Respond to ALL reviews within 24h","Positive: thank + mention VIDS. Negative: acknowledge + offer offline resolution",PENDING),
    (20,"Weekly GMB post schedule","Every Monday: vastu tip. Every Thursday: client success or educational. Rotate topics.",PENDING),
]
dv_g=DataValidation(type="list",formula1='"Done,In Progress,Pending"',allow_blank=False)
ws_g.add_data_validation(dv_g)
for ri,(n,task,what,sta) in enumerate(gmb,4):
    ws_g.row_dimensions[ri].height=44
    bg=LT_GREEN if sta==DONE else LT_RED
    for ci,val in enumerate([n,task,what,sta,""],1):
        c=ws_g.cell(ri,ci,val); c.fill=fill(bg if ci>1 else GRAY); c.border=bdr()
        c.font=Font(bold=(ci==2),color="111111",size=10,name="Calibri")
        c.alignment=ctr() if ci in(1,4,5) else lft()
    dv_g.add(ws_g.cell(ri,4))
ws_g.freeze_panes="A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: CONTENT CALENDAR
# ══════════════════════════════════════════════════════════════════════════════
ws_cc = wb.create_sheet("Content Calendar")
ws_cc.sheet_view.showGridLines = False
for col,w in zip("ABCDEFG",[5,28,38,16,16,28,20]):
    ws_cc.column_dimensions[col].width=w

ws_cc.merge_cells("A1:G1")
c=ws_cc["A1"]; c.value="Content Calendar — Google News + RSS + Blog + GMB Posts"
c.fill=fill("1565C0"); c.font=font(True,size=14); c.alignment=ctr()
ws_cc.row_dimensions[1].height=40

ws_cc.merge_cells("A2:G2")
c=ws_cc["A2"]; c.value="2 blog posts/week = Google News eligibility. 1 GMB post/week = map rank boost. All feed RSS."
c.fill=fill(GOLD); c.font=font(size=11); c.alignment=ctr()
ws_cc.row_dimensions[2].height=24

for ci,h in enumerate(["#","Content Type","Title / Topic","Platform","Status","Notes","Publish Date"],1):
    c=ws_cc.cell(3,ci,h); c.fill=fill(NAVY); c.font=font(True); c.alignment=ctr(); c.border=bdr()
ws_cc.row_dimensions[3].height=26

calendar=[
    (1,"Blog Post","5 Vastu Mistakes in Bangalore Apartments That Cost You Peace","Blog + Google News + RSS",PENDING,"High local intent — link to room pages"),
    (2,"GMB Post","Vastu Tip: Why your refrigerator should never be in the NE corner","GMB",PENDING,"Link to /vastu-for-kitchen"),
    (3,"Blog Post","Vastu for 2BHK Apartments: Room-by-Room Guide","Blog + Google News",PENDING,"Link to bedroom, kitchen, bathroom pages"),
    (4,"GMB Post","Client Result: Corrected master bedroom zone — better sleep in 7 days","GMB",PENDING,"Social proof post"),
    (5,"Blog Post","South-Facing House Vastu: Why It Is Not Inauspicious (The Truth)","Blog + Google News",PENDING,"High volume search — controversial angle = shares"),
    (6,"GMB Post","Vastu for home office — Work from home Vastu tips","GMB",PENDING,"Link to /vastu-for-study-room"),
    (7,"Blog Post","VIDS System Explained: How 16-Zone Vastu Analysis Works","Blog + Google News",PENDING,"Differentiator content — brand building"),
    (8,"GMB Post","Now serving all Whitefield and Sarjapur Road apartments — book a consultation","GMB",PENDING,"Locality targeting"),
    (9,"Blog Post","Vastu Before Buying a Flat: 10 Things to Check","Blog + Google News",PENDING,"Apartment buyer intent — link to apartment pages"),
    (10,"GMB Post","Vastu for Pooja Room: NE is the sacred zone for maximum divine grace","GMB",PENDING,"Link to /vastu-for-pooja-room"),
    (11,"Blog Post","Vastu for IT Professionals in Bangalore: Office and Home Alignment","Blog + Google News",PENDING,"IT corridor targeting"),
    (12,"GMB Post","Online Vastu Consultations Available Worldwide — Book via WhatsApp","GMB",PENDING,"International + online offer"),
    (13,"Blog Post","Vastu for Factory and Industrial Spaces: Zone Rules for Productivity","Blog + Google News",PENDING,"Industrial segment — new revenue stream"),
    (14,"GMB Post","Vastu Tip: The one correction that improves sleep immediately","GMB",PENDING,"Head direction — bedroom page link"),
    (15,"Blog Post","Vardhini Vastu Case Study: Zero-Demolition Correction in Whitefield Apartment","Blog + Google News",PENDING,"E-E-A-T and social proof"),
    (16,"Quora Answer","Answered: Best vastu consultant in Bangalore?","Quora + Off-page",PENDING,"2 answers/week — link to site"),
    (17,"Blog Post","Vastu for Devanahalli and Airport Road Homes: Luxury Zone Guide","Blog + Google News",PENDING,"New locality targeting"),
    (18,"X/Twitter Post","Thread: 7 vastu rules for the bedroom [with images]","X/Twitter + Grok",PENDING,"3 posts/week for Grok AI training"),
    (19,"Blog Post","Vastu for NRIs: How to Check Vastu of Your India Home Remotely","Blog + Google News",PENDING,"International + online consultation"),
    (20,"GMB Post","Now covering Devanahalli, Airport Road, and Hebbal — call to book","GMB",PENDING,"Extended north Bangalore locality push"),
]
dv_cc=DataValidation(type="list",formula1='"Done,In Progress,Pending,Scheduled"',allow_blank=False)
ws_cc.add_data_validation(dv_cc)
for ri,(n,ctype,title,platform,sta,notes) in enumerate(calendar,4):
    ws_cc.row_dimensions[ri].height=38
    bg=LT_GREEN if sta==DONE else LT_RED
    for ci,val in enumerate([n,ctype,title,platform,sta,notes,""],1):
        c=ws_cc.cell(ri,ci,val); c.fill=fill(bg if ci>1 else GRAY); c.border=bdr()
        c.font=Font(bold=(ci in(2,3)),color="111111",size=10,name="Calibri")
        c.alignment=ctr() if ci in(1,5) else lft()
    dv_cc.add(ws_cc.cell(ri,5))
ws_cc.freeze_panes="A4"

# ── Save ──────────────────────────────────────────────────────────────────────
out = "C:/Users/raghu/VV/VardhiniVastu_SEO_Tracker_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
total=len(tasks); done_n=sum(1 for t in tasks if t[3]==DONE)
print(f"Tasks: {total} total | {done_n} done | {total-done_n} pending")
print(f"Localities: 50 total (20 live + 30 extended)")
